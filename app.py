"""
Lutron Designer File Viewer / Editor
Supports .pl files (ZIP archive) and extracted folders.
"""

import os
import re
import uuid
import sqlite3
import zipfile
import shutil
import tempfile
import threading
import webbrowser
import tkinter as tk
from tkinter import filedialog
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import Flask, jsonify, request, send_from_directory, send_file, Response

try:
    import pyodbc
    PYODBC_OK = True
except ImportError:
    PYODBC_OK = False

app = Flask(__name__, static_folder="static", template_folder="static")

state = {
    "pl_path":       None,   # original .pl file path (None if folder was selected)
    "work_dir":      None,   # temp dir extracted from .pl, or the folder itself
    "db_name":       None,   # SQL Server temp DB name
    "lut_path":      None,   # path of extracted .lut
    "dirty":         False,  # unsaved changes exist
    "template_id":   None,   # active template ID (filename in Templates/)
    "tmpl_zip_name": None,   # same as template_id (kept for clarity in save)
    "undo_stack":    [],     # [{label, sqls, redo_sqls}, ...]  max 50
    "redo_stack":    [],
}

def _detect_sql_instance():
    """Auto-detect the installed LUTRON SQL Server instance from the registry.
    If multiple LUTRON instances exist, prefer the one with the highest version number.
    Falls back to LUTRON2022 if none found."""
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\Microsoft\Microsoft SQL Server\Instance Names\SQL",
        )
        lutron_names = []
        i = 0
        while True:
            try:
                name, _, _ = winreg.EnumValue(key, i)
                i += 1
                if "LUTRON" in name.upper():
                    lutron_names.append(name)
            except OSError:
                break
        winreg.CloseKey(key)
        if lutron_names:
            # Pick the instance with the highest trailing year/number (e.g. LUTRON2022 > LUTRON2019)
            lutron_names.sort(key=lambda n: int(''.join(filter(str.isdigit, n)) or '0'), reverse=True)
            return rf".\{lutron_names[0]}"
    except OSError:
        pass
    return r".\LUTRON2022"

SQL_INSTANCE = _detect_sql_instance()


def diagnose_sql():
    """Return a human-readable diagnostic string for SQL Server connectivity."""
    lines = []

    if not PYODBC_OK:
        lines.append("❌ pyodbc 未インストール (pip install pyodbc)")
        return "\n".join(lines)

    # Installed ODBC drivers
    available = pyodbc.drivers()
    sql_drivers = [d for d in available if "SQL" in d.upper()]
    if sql_drivers:
        lines.append(f"✅ ODBCドライバー: {', '.join(sql_drivers)}")
    else:
        lines.append("❌ SQL Server用ODBCドライバーなし → msodbcsql17.msi をインストールしてください")

    # Registry: enumerate installed SQL Server instances
    lutron_instances = []
    try:
        import winreg
        key = winreg.OpenKey(
            winreg.HKEY_LOCAL_MACHINE,
            r"SOFTWARE\Microsoft\Microsoft SQL Server\Instance Names\SQL",
        )
        i = 0
        while True:
            try:
                name, _, _ = winreg.EnumValue(key, i)
                i += 1
                if "LUTRON" in name.upper():
                    lutron_instances.append(name)
            except OSError:
                break
        winreg.CloseKey(key)
    except OSError:
        pass

    if lutron_instances:
        lines.append(f"✅ LUTRONインスタンス(レジストリ): {', '.join(lutron_instances)}")
    else:
        lines.append("❌ レジストリにLUTRONインスタンスなし → Lutron Designerが正常にインストールされていない可能性")

    # Check configured instance matches registry
    configured = SQL_INSTANCE.lstrip(".\\").lstrip("./")
    lines.append(f"ℹ️  接続先インスタンス: {SQL_INSTANCE}")
    if lutron_instances and configured.upper() not in [n.upper() for n in lutron_instances]:
        lines.append(f"⚠️  '{configured}' はレジストリ未検出。実在するインスタンス: {', '.join(lutron_instances)}")

    # Live connection test per driver
    lines.append("--- 接続テスト ---")
    for drv in ["ODBC Driver 17 for SQL Server", "SQL Server Native Client 11.0", "SQL Server"]:
        if drv not in available:
            lines.append(f"  スキップ (未インストール): {drv}")
            continue
        try:
            c = pyodbc.connect(
                f"DRIVER={{{drv}}};SERVER={SQL_INSTANCE};DATABASE=master;Trusted_Connection=yes;",
                timeout=5,
            )
            c.close()
            lines.append(f"  ✅ {drv}")
        except pyodbc.Error as e:
            lines.append(f"  ❌ {drv}: {e}")
    return "\n".join(lines)


def _run_sql_admin(sql, timeout=120):
    """Run a DDL statement (RESTORE/BACKUP/DROP) against master with autocommit."""
    conn = None
    for drv in ["ODBC Driver 17 for SQL Server", "SQL Server Native Client 11.0", "SQL Server"]:
        try:
            conn = pyodbc.connect(
                f"DRIVER={{{drv}}};SERVER={SQL_INSTANCE};DATABASE=master;Trusted_Connection=yes;",
                autocommit=True,
                timeout=30,
            )
            break
        except pyodbc.Error:
            continue
    if conn is None:
        raise RuntimeError("SQL Server に接続できません\n\n" + diagnose_sql())
    conn.timeout = timeout
    cur = conn.cursor()
    try:
        cur.execute(sql)
    except pyodbc.Error as e:
        sqlstate = str(e.args[0]) if e.args else ""
        if not sqlstate.startswith("01"):
            conn.close()
            raise RuntimeError(str(e))
    # Consume all remaining result sets (RESTORE/BACKUP sends multiple info messages).
    # This ensures SQL Server finishes before we close the connection.
    while True:
        try:
            if not cur.nextset():
                break
        except pyodbc.Error as e:
            sqlstate = str(e.args[0]) if e.args else ""
            if not sqlstate.startswith("01"):
                conn.close()
                raise RuntimeError(str(e))
        except StopIteration:
            break
    conn.close()


# ─────────────────────────────────────────────
# Helpers – SQL Server
# ─────────────────────────────────────────────

def sql_conn():
    for drv in ["ODBC Driver 17 for SQL Server", "SQL Server Native Client 11.0", "SQL Server"]:
        try:
            return pyodbc.connect(
                f"DRIVER={{{drv}}};SERVER={SQL_INSTANCE};DATABASE=master;Trusted_Connection=yes;",
                timeout=10
            )
        except pyodbc.Error:
            continue
    raise RuntimeError("SQL Server に接続できません\n\n" + diagnose_sql())


def q(sql, params=()):
    """SELECT against current temp DB, return list of dicts."""
    conn = sql_conn()
    cur = conn.cursor()
    cur.execute(f"USE [{state['db_name']}]")
    cur.execute(sql, params)
    cols = [d[0] for d in cur.description]
    rows = [dict(zip(cols, r)) for r in cur.fetchall()]
    conn.close()
    return rows


def execute_sql(sql, params=()):
    """DML (INSERT/UPDATE/DELETE) against current temp DB."""
    conn = sql_conn()
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"USE [{state['db_name']}]")
    cur.execute(sql, params)
    conn.commit()
    conn.close()
    state["dirty"] = True


def _alloc_id(cur):
    """Allocate a unique object ID from tblNextObjectID (Lutron's ID sequencer).
    Returns the allocated ID (int). Must be called inside a transaction."""
    cur.execute("SELECT NextObjectID FROM tblNextObjectID WITH (UPDLOCK)")
    row = cur.fetchone()
    new_id = (row[0] if row else 9000) + 1
    if row:
        cur.execute("UPDATE tblNextObjectID SET NextObjectID = ?", (new_id,))
    else:
        cur.execute("INSERT INTO tblNextObjectID (NextObjectID) VALUES (?)", (new_id,))
    return new_id


def _alloc_and_insert(table, pk_col, row):
    """Allocate a new ID from tblNextObjectID, INSERT row dict into table.
    pk_col: name of the primary key column (will be set to the allocated ID).
    row: dict of {column: value} pairs (must NOT include pk_col).
    Returns the new integer ID."""
    conn = sql_conn()
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"USE [{state['db_name']}]")
    new_id = _alloc_id(cur)
    cols = [pk_col] + list(row.keys())
    vals = [new_id] + list(row.values())
    col_str = ", ".join(cols)
    ph_str = ", ".join("?" * len(vals))
    cur.execute(f"INSERT INTO {table} ({col_str}) VALUES ({ph_str})", vals)
    conn.commit()
    conn.close()
    state["dirty"] = True
    return new_id


def execute_sqls(statements):
    """Execute multiple DML statements in a single transaction."""
    conn = sql_conn()
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute(f"USE [{state['db_name']}]")
    for sql, params in statements:
        cur.execute(sql, params)
    conn.commit()
    conn.close()
    state["dirty"] = True


def push_undo(undo_sqls, redo_sqls, label=""):
    """Push a reversible operation onto the undo stack.
    undo_sqls / redo_sqls: list of (sql_string, params_tuple) pairs.
    Clears the redo stack on new push."""
    state["undo_stack"].append({"label": label, "sqls": undo_sqls, "redo_sqls": redo_sqls})
    state["redo_stack"].clear()
    if len(state["undo_stack"]) > 50:
        state["undo_stack"].pop(0)


# LoadType → ControlType mapping
_CCO_MAINTAINED = {25, 37}
_CCO_PULSED     = {26}
_HVAC           = {48, 135}
_FAN            = {7, 46, 127, 128}
_SHADE          = {36, 53, 129, 130, 149}
_VENETIAN       = {49}
_SWITCHED       = {4, 17, 18, 20, 21, 22, 23, 24, 32, 33, 35, 40, 54, 111, 118, 133, 134}

def load_type_to_control_type(lt):
    if lt in _CCO_MAINTAINED: return 7
    if lt in _CCO_PULSED:     return 8
    if lt in _HVAC:           return 9
    if lt in _FAN:            return 6
    if lt in _SHADE:          return 5
    if lt in _VENETIAN:       return 10
    if lt in _SWITCHED:       return 2
    return 1  # Dimmed


def drop_db(db_name):
    try:
        drop_sql = (
            f"IF DB_ID(N'{db_name}') IS NOT NULL BEGIN "
            f"ALTER DATABASE [{db_name}] SET SINGLE_USER WITH ROLLBACK IMMEDIATE; "
            f"DROP DATABASE [{db_name}]; END"
        )
        _run_sql_admin(drop_sql, timeout=30)
    except Exception:
        pass
    lut_dest = os.path.join(r"C:\ProgramData\Lutron", f"{db_name}.lut")
    try:
        if os.path.exists(lut_dest):
            os.remove(lut_dest)
    except OSError:
        pass


def restore_lut(lut_path, db_name):
    mdf = os.path.join(r"C:\ProgramData\Lutron", f"{db_name}.mdf")
    ldf = os.path.join(r"C:\ProgramData\Lutron", f"{db_name}_log.ldf")
    lut_dest = os.path.join(r"C:\ProgramData\Lutron", f"{db_name}.lut")

    drop_sql = (
        f"IF DB_ID(N'{db_name}') IS NOT NULL BEGIN "
        f"ALTER DATABASE [{db_name}] SET SINGLE_USER WITH ROLLBACK IMMEDIATE; "
        f"DROP DATABASE [{db_name}]; END"
    )
    try:
        _run_sql_admin(drop_sql, timeout=30)
    except Exception:
        pass

    if os.path.exists(lut_dest):
        try:
            os.remove(lut_dest)
        except OSError:
            lut_dest = os.path.join(r"C:\ProgramData\Lutron",
                                    f"{db_name}_{uuid.uuid4().hex[:8]}.lut")

    shutil.copy2(lut_path, lut_dest)

    sql = (
        f"RESTORE DATABASE [{db_name}] FROM DISK = N'{lut_dest}' WITH "
        f"MOVE N'Project' TO N'{mdf}', "
        f"MOVE N'Project_log' TO N'{ldf}', RECOVERY, REPLACE"
    )
    _run_sql_admin(sql, timeout=120)


# ─────────────────────────────────────────────
# Helpers – .pl / folder loading
# ─────────────────────────────────────────────

def sqlite_rows(path, query, params=()):
    con = sqlite3.connect(path)
    con.row_factory = sqlite3.Row
    cur = con.execute(query, params)
    rows = [dict(r) for r in cur.fetchall()]
    con.close()
    return rows


def extract_pl_to_temp(pl_path):
    """Extract .pl (ZIP) to a fresh temp dir and return its path."""
    if not zipfile.is_zipfile(pl_path):
        raise ValueError(f"このファイルはZIP形式の .pl ファイルではありません: {os.path.basename(pl_path)}")
    out = tempfile.mkdtemp(prefix="ld_pl_")
    with zipfile.ZipFile(pl_path) as z:
        z.extractall(out)
    return out


def list_template_zips(folder):
    """Return sorted list of template ZIP filenames found in Templates/."""
    tmpl_dir = os.path.join(folder, "Templates")
    if not os.path.isdir(tmpl_dir):
        return []
    result = []
    for name in sorted(os.listdir(tmpl_dir)):
        fp = os.path.join(tmpl_dir, name)
        if not os.path.isfile(fp):
            continue
        try:
            with zipfile.ZipFile(fp) as z:
                if any(zi.filename.endswith(".lut") for zi in z.infolist()):
                    result.append(name)
        except Exception:
            pass
    return result


def find_lut_for_template(folder, template_id=None):
    """
    Extract .lut for the given template_id (or first available).
    Returns (lut_path, tmpl_zip_name) or (None, None).
    """
    tmpl_dir = os.path.join(folder, "Templates")
    if not os.path.isdir(tmpl_dir):
        return None, None

    candidates = list_template_zips(folder)
    if not candidates:
        return None, None

    if template_id and template_id in candidates:
        candidates = [template_id] + [c for c in candidates if c != template_id]

    for name in candidates:
        fp = os.path.join(tmpl_dir, name)
        try:
            with zipfile.ZipFile(fp) as z:
                for zi in z.infolist():
                    if zi.filename.endswith(".lut"):
                        out_dir = tempfile.mkdtemp(prefix="ld_lut_")
                        z.extract(zi, out_dir)
                        return os.path.join(out_dir, zi.filename), name
        except Exception:
            pass
    return None, None


def load_project(work_dir, template_id=None):
    """Restore the SQL Server DB from work_dir and update state."""
    if state["db_name"]:
        drop_db(state["db_name"])
    state["db_name"] = None
    state["lut_path"] = None
    state["template_id"] = None
    state["tmpl_zip_name"] = None
    state["dirty"] = False
    state["undo_stack"] = []
    state["redo_stack"] = []

    lut, tmpl_zip_name = find_lut_for_template(work_dir, template_id)
    if lut and PYODBC_OK:
        db_name = "LDViewer_Temp"
        restore_lut(lut, db_name)
        state["db_name"] = db_name
        state["lut_path"] = lut
        state["template_id"] = tmpl_zip_name
        state["tmpl_zip_name"] = tmpl_zip_name
        # Repair: remove invalid value=2 written to Room Property PT rows by old UI code.
        # LD encoding: 65280=Unaffected, 0=second state(Disabled/Off), 1=first state(Enabled/On).
        # value=2 does not exist in LD's enum and causes IndexOutOfRangeException.
        try:
            execute_sql("""
                DELETE FROM tblAssignmentCommandParameter
                WHERE ParameterValue = 2
                  AND ParentId IN (
                      SELECT PresetAssignmentID FROM tblPresetAssignment
                      WHERE AssignableObjectType = 400
                  )
            """)
        except Exception as _e:
            app.logger.warning(f"[load_project] Room Property invalid value cleanup failed: {_e}")


def backup_lut_from_db(db_name, dest_lut_path):
    """Backup the current SQL Server DB back to a .lut file."""
    sql = f"BACKUP DATABASE [{db_name}] TO DISK = N'{dest_lut_path}' WITH FORMAT, INIT"
    _run_sql_admin(sql, timeout=120)


def save_back_to_pl():
    """
    Backup current DB → .lut, repack into the Template ZIP,
    then repack everything back into the .pl file.
    """
    if not state["pl_path"] or not state["db_name"]:
        raise RuntimeError("保存対象がありません")
    if not state["tmpl_zip_name"]:
        raise RuntimeError("テンプレート情報がありません")

    work_dir = state["work_dir"]
    tmpl_dir = os.path.join(work_dir, "Templates")
    tmpl_zip_name = state["tmpl_zip_name"]
    tmpl_zip_path = os.path.join(tmpl_dir, tmpl_zip_name)

    # Read all entries from the original template ZIP (preserve non-.lut files)
    with zipfile.ZipFile(tmpl_zip_path) as z:
        lut_entry = next(zi.filename for zi in z.infolist() if zi.filename.endswith(".lut"))
        other_entries = [
            (zi, z.read(zi.filename))
            for zi in z.infolist()
            if not zi.filename.endswith(".lut")
        ]
    app.logger.info(f"[save] Template ZIP: lut={lut_entry}, other files={len(other_entries)}")

    # Backup DB → C:\ProgramData\Lutron\ using a unique name to avoid file lock issues
    tmp_lut = os.path.join(r"C:\ProgramData\Lutron", f"LDViewer_bak_{uuid.uuid4().hex[:8]}.lut")

    app.logger.info(f"[save] Backing up DB to {tmp_lut}")
    backup_lut_from_db(state["db_name"], tmp_lut)

    if not os.path.exists(tmp_lut):
        raise RuntimeError(f"バックアップ失敗: {tmp_lut} が作成されませんでした")
    app.logger.info(f"[save] Backup size: {os.path.getsize(tmp_lut):,} bytes")

    # Repack: preserve all original files, replace only the .lut
    tmp_zip = tmpl_zip_path + ".new"
    app.logger.info(f"[save] Repacking template ZIP")
    with zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as zout:
        for zi, data in other_entries:
            zout.writestr(zi, data)
        zout.write(tmp_lut, lut_entry)
    os.replace(tmp_zip, tmpl_zip_path)
    app.logger.info(f"[save] Template ZIP updated: {os.path.getsize(tmpl_zip_path):,} bytes")
    try:
        os.remove(tmp_lut)
    except OSError:
        pass

    # Write new .pl to a LOCAL temp directory first, then copy to network drive
    pl_path = state["pl_path"]
    tmp_pl_dir = tempfile.mkdtemp(prefix="ld_save_")
    tmp_pl = os.path.join(tmp_pl_dir, os.path.basename(pl_path))
    try:
        app.logger.info(f"[save] Building new .pl at {tmp_pl}")
        with zipfile.ZipFile(tmp_pl, "w", zipfile.ZIP_DEFLATED) as zout:
            for root, dirs, files in os.walk(work_dir):
                for fname in files:
                    abs_path = os.path.join(root, fname)
                    arc_name = os.path.relpath(abs_path, work_dir).replace("\\", "/")
                    zout.write(abs_path, arc_name)
            # Preserve empty Systems/ directory entry
            sys_dir = os.path.join(work_dir, "Systems")
            if os.path.isdir(sys_dir) and not os.listdir(sys_dir):
                if hasattr(zout, "mkdir"):
                    zout.mkdir("Systems/")

        app.logger.info(f"[save] New .pl size: {os.path.getsize(tmp_pl):,} bytes")
        app.logger.info(f"[save] Copying to {pl_path}")
        shutil.copy2(tmp_pl, pl_path)
        app.logger.info(f"[save] Done. Final .pl size: {os.path.getsize(pl_path):,} bytes")
    finally:
        shutil.rmtree(tmp_pl_dir, ignore_errors=True)

    state["dirty"] = False


# ─────────────────────────────────────────────
# API Routes
# ─────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory("static", "index.html")


@app.route("/api/diagnostics")
def get_diagnostics():
    """Return SQL Server connectivity diagnostics."""
    return jsonify({"report": diagnose_sql()})


@app.route("/api/debug/roomprop-assignments")
def debug_roomprop_assignments():
    """Dump all Room Property assignments with scene name and human-readable PT values."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    # PT → property name mapping (for display only)
    PT_NAMES = {76: "RoomState(152)", 77: "Presence(149)", 78: "FGE(154)",
                79: "Turndown(155)", 80: "DND(151)", 81: "MUR(153)", 84: "Auto(157)"}
    rows = q("""
        SELECT pa.PresetAssignmentID, pa.ParentID, pa.ParentType,
               pa.AssignableObjectType, pa.AssignmentCommandType, pa.AssignmentCommandGroup,
               p.Name AS SceneName, p.PresetType,
               ar.Name AS AreaName
        FROM tblPresetAssignment pa
        LEFT JOIN tblPreset p ON p.PresetID = pa.ParentID
        LEFT JOIN tblArea ar ON ar.AreaID = p.ParentID
        WHERE pa.AssignableObjectType = 400
        ORDER BY pa.ParentID, pa.PresetAssignmentID
    """)
    result = []
    for row in rows:
        aid = row["PresetAssignmentID"]
        params = q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=? ORDER BY ParameterType", (aid,))
        pt_map = {p["ParameterType"]: p["ParameterValue"] for p in params}
        decoded = {}
        for pt, val in pt_map.items():
            name = PT_NAMES.get(pt, f"PT{pt}")
            decoded[name] = "Unaffected(65280)" if val == 65280 else str(val)
        result.append({
            "PresetAssignmentID": aid,
            "ParentID": row["ParentID"],
            "ParentType": row["ParentType"],
            "SceneName": row["SceneName"],
            "PresetType": row["PresetType"],
            "AreaName": row["AreaName"],
            "raw_params": pt_map,
            "decoded": decoded,
        })
    return jsonify(result)


@app.route("/api/open", methods=["POST"])
def open_file():
    """Open a .pl file or a folder via web upload."""
    mode = request.form.get("mode", "pl")  # "pl" or "folder"

    if mode == "pl":
        if "file" not in request.files:
            return jsonify({"error": "ファイルがアップロードされていません"}), 400
        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "ファイルが選択されていません"}), 400

        if state["work_dir"] and state["pl_path"]:
            shutil.rmtree(state["work_dir"], ignore_errors=True)

        temp_pl_fd, temp_pl_path = tempfile.mkstemp(suffix=".pl", prefix="ld_upload_")
        os.close(temp_pl_fd)
        file.save(temp_pl_path)

        try:
            work_dir = extract_pl_to_temp(temp_pl_path)
        except (ValueError, zipfile.BadZipFile) as e:
            return jsonify({"error": str(e)}), 400
            
        state["pl_path"] = temp_pl_path
        state["work_dir"] = work_dir
        path_name = request.form.get("filename", file.filename)
        state["original_filename"] = path_name

    else:  # folder
        files = request.files.getlist("files")
        paths = request.form.getlist("paths")
        if not files or not paths:
            return jsonify({"error": "ファイルがアップロードされていません"}), 400
            
        work_dir = tempfile.mkdtemp(prefix="ld_upload_dir_")
        
        for f, rel_path in zip(files, paths):
            parts = rel_path.replace('\\', '/').split('/')
            if len(parts) > 1:
                rel_val = "/".join(parts[1:])
            else:
                rel_val = parts[0]
                
            target_path = os.path.join(work_dir, rel_val)
            os.makedirs(os.path.dirname(target_path), exist_ok=True)
            f.save(target_path)
            
        state["pl_path"] = None
        state["work_dir"] = work_dir
        path_name = parts[0] if parts else "Uploaded Folder"
        state["original_filename"] = path_name

    sqlite_path = os.path.join(state["work_dir"], "PlaceCache.sqlite")
    if not os.path.exists(sqlite_path):
        return jsonify({"error": "PlaceCache.sqlite が見つかりません"}), 400

    try:
        load_project(state["work_dir"])
    except Exception as e:
        err_str = str(e)
        if "3169" in err_str or "version 16" in err_str or "incompatible with this server" in err_str:
            return jsonify({"error": (
                "SQL Server バージョン不一致エラー: このバックアップは SQL Server 2022 (v16) で作成されましたが、"
                "このPCには古いバージョン (SQL Server 2019 など) がインストールされています。\n"
                "解決方法: このPCに SQL Server 2022 をインストールしてください。"
            )}), 500
        return jsonify({"error": f"SQL Server復元失敗: {e}"}), 500

    templates = list_template_zips(state["work_dir"])

    return jsonify({
        "path": path_name,
        "name": os.path.basename(path_name),
        "is_pl": mode == "pl",
        "has_template": state["db_name"] is not None,
        "templates": templates,
        "active_template": state["template_id"],
    })


@app.route("/api/switch-template", methods=["POST"])
def switch_template():
    """Switch to a different template and reload the DB."""
    if not state["work_dir"]:
        return jsonify({"error": "未選択"}), 400
    template_id = request.json.get("template_id")
    if not template_id:
        return jsonify({"error": "template_id 必須"}), 400
    try:
        load_project(state["work_dir"], template_id)
    except Exception as e:
        return jsonify({"error": f"テンプレート切替失敗: {e}"}), 500
    return jsonify({"ok": True, "active_template": state["template_id"]})


@app.route("/api/overview")
def overview():
    if not state["work_dir"]:
        return jsonify({"error": "未選択"}), 400
    sqlite_path = os.path.join(state["work_dir"], "PlaceCache.sqlite")
    place = sqlite_rows(sqlite_path, "SELECT * FROM tblPlace LIMIT 1")
    templates = sqlite_rows(sqlite_path, "SELECT id, name, modifiedDate, systemCount, projectDatabaseArtifactId FROM tblTemplate")
    systems = sqlite_rows(sqlite_path, "SELECT * FROM tblTemplatedSystem")
    addr = sqlite_rows(sqlite_path, "SELECT * FROM tblAddressDetails LIMIT 1")
    fs_templates = list_template_zips(state["work_dir"])
    return jsonify({
        "place": place[0] if place else {},
        "templates": templates,
        "fs_templates": fs_templates,
        "active_template": state["template_id"],
        "systems": systems,
        "address": addr[0] if addr else {},
        "dirty": state["dirty"],
        "is_pl": state["pl_path"] is not None,
    })


@app.route("/api/areas")
def areas():
    if not state["db_name"]:
        return jsonify([])
    rows = q("""
        SELECT AreaID, Name, ParentID, HierarchyLevel, IsLeaf, AreaType
        FROM tblArea
        WHERE HierarchyLevel >= 2
        ORDER BY HierarchyLevel, ParentID, SortOrder, Name
    """)
    return jsonify(rows)


@app.route("/api/area/<int:area_id>")
def area_detail(area_id):
    if not state["db_name"]:
        return jsonify({})

    zones = q("""
        SELECT z.ZoneID, z.Name, z.ZoneNumber as LoadNumber,
               z.ControlType, z.ZoneLayer, z.ZoneDescription,
               sl.LoadType, lt.Description as LoadTypeDesc
        FROM tblZone z
        LEFT JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
        LEFT JOIN lstLoadType lt ON lt.LoadTypeID = sl.LoadType
        WHERE z.ParentID = ?
        ORDER BY z.ZoneNumber
    """, (area_id,))

    stations = q("""
        SELECT cs.ControlStationID, cs.Name,
               csd.ControlStationDeviceID, csd.Name as DeviceName, csd.ModelInfoID
        FROM tblControlStation cs
        LEFT JOIN tblControlStationDevice csd ON csd.ParentControlStationID = cs.ControlStationID
        WHERE cs.ParentId = ? AND cs.ParentType = 2
        ORDER BY cs.SortOrder, cs.Name
    """, (area_id,))

    station_ids = list({s["ControlStationDeviceID"] for s in stations if s["ControlStationDeviceID"]})
    buttons = []
    if station_ids:
        ph = ",".join("?" * len(station_ids))
        buttons = q(f"""
            SELECT kb.ButtonID, kb.ButtonNumber, kb.Name as ButtonName,
                   kb.ParentDeviceID,
                   pm.Name as ProgModel, pm.ControlType as ProgControlType,
                   pm.PresetID
            FROM tblKeypadButton kb
            LEFT JOIN tblProgrammingModel pm ON kb.ProgrammingModelID = pm.ProgrammingModelID
            WHERE kb.ParentDeviceID IN ({ph})
            ORDER BY kb.ParentDeviceID, kb.ButtonNumber
        """, station_ids)

    scenes = q("""
        SELECT s.SceneID, s.Name, s.Number, sc.SceneControllerID
        FROM tblSceneController sc
        JOIN tblScene s ON s.ParentSceneControllerID = sc.SceneControllerID
        WHERE sc.ParentID = ? AND sc.ParentType = 2
        ORDER BY s.Number
    """, (area_id,))

    return jsonify({
        "zones": zones,
        "stations": stations,
        "buttons": buttons,
        "scenes": scenes,
    })


@app.route("/api/shared-scenes/export")
def export_shared_scenes():
    """Export all shared scenes (all areas) with zone assignments as xlsx or csv."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    fmt = request.args.get("format", "xlsx").lower()

    rows = q("""
        SELECT
            a.Name  AS AreaName,
            p.Name  AS SceneName,
            p.SortOrder,
            z.Name  AS ZoneName,
            acp3.ParameterValue AS Level,
            acp1.ParameterValue AS FadeRaw,
            acp2.ParameterValue AS DelayRaw
        FROM tblArea a
        JOIN tblPreset p
            ON p.ParentID = a.AreaID AND p.PresetType = 3 AND p.ParentType = 2
        LEFT JOIN tblPresetAssignment pa
            ON pa.ParentID = p.PresetID AND pa.AssignableObjectType = 15
        LEFT JOIN tblZone z
            ON z.ZoneID = pa.AssignableObjectID
        LEFT JOIN tblAssignmentCommandParameter acp3
            ON acp3.ParentId = pa.PresetAssignmentID AND acp3.ParameterType = 3
        LEFT JOIN tblAssignmentCommandParameter acp1
            ON acp1.ParentId = pa.PresetAssignmentID AND acp1.ParameterType = 1
        LEFT JOIN tblAssignmentCommandParameter acp2
            ON acp2.ParentId = pa.PresetAssignmentID AND acp2.ParameterType = 2
        WHERE a.HierarchyLevel >= 2
        ORDER BY a.Name, p.SortOrder, p.Name, z.Name
    """)

    headers = ["Area", "Scene", "Zone", "Level", "Fade(s)", "Delay(s)"]
    data_rows = []
    for r in rows:
        fade  = round(r["FadeRaw"]  / 4, 2) if r["FadeRaw"]  is not None else ""
        delay = round(r["DelayRaw"] / 4, 2) if r["DelayRaw"] is not None else ""
        level = r["Level"] if r["Level"] is not None else ""
        data_rows.append([r["AreaName"], r["SceneName"],
                          r["ZoneName"] or "", level, fade, delay])

    if fmt == "csv":
        buf = io.StringIO()
        csv.writer(buf).writerow(headers)
        csv.writer(buf).writerows(data_rows)
        output = buf.getvalue().encode("utf-8-sig")
        return Response(output, mimetype="text/csv",
                        headers={"Content-Disposition":
                                 "attachment; filename=shared_scenes.csv"})

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shared Scenes"
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
    for row in data_rows:
        ws.append(row)
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 40)
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":
                             "attachment; filename=shared_scenes.xlsx"})


@app.route("/api/shared-scenes/import", methods=["POST"])
def import_shared_scenes():
    """Import shared scenes from xlsx or csv.
    Columns: Area, Scene, Zone, Level, Fade(s), Delay(s)
    Creates missing scenes; upserts zone assignments.
    """
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    if "file" not in request.files:
        return jsonify({"error": "ファイルが必要です"}), 400
    f = request.files["file"]
    fname = (f.filename or "").lower()

    # Parse file into list of dicts
    file_rows = []
    try:
        if fname.endswith(".csv"):
            content = f.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            file_rows = list(reader)
        elif fname.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(f.read()), read_only=True)
            ws = wb.active
            hdrs = [str(c.value or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
            for ws_row in ws.iter_rows(min_row=2, values_only=True):
                file_rows.append(dict(zip(hdrs, ws_row)))
        else:
            return jsonify({"error": "xlsx または csv ファイルのみ対応"}), 400
    except Exception as e:
        return jsonify({"error": f"ファイル読み込みエラー: {e}"}), 400

    # Cache areas
    areas = q("SELECT AreaID, Name FROM tblArea WHERE HierarchyLevel >= 2")
    area_map = {a["Name"]: a["AreaID"] for a in areas}

    created_scenes = 0
    updated_assignments = 0
    errors = []
    scene_cache = {}   # (area_id, scene_name) → preset_id

    for i, row in enumerate(file_rows, start=2):
        area_name  = str(row.get("Area")    or "").strip()
        scene_name = str(row.get("Scene")   or "").strip()
        zone_name  = str(row.get("Zone")    or "").strip()
        level_str  = str(row.get("Level")   or "").strip()
        fade_str   = str(row.get("Fade(s)") or "2").strip()
        delay_str  = str(row.get("Delay(s)") or "0").strip()

        if not area_name or not scene_name:
            continue

        area_id = area_map.get(area_name)
        if area_id is None:
            errors.append(f"行{i}: エリア '{area_name}' が見つかりません")
            continue

        # Find or create scene
        cache_key = (area_id, scene_name)
        if cache_key not in scene_cache:
            ex = q("""SELECT PresetID FROM tblPreset
                      WHERE PresetType=3 AND ParentType=2 AND ParentID=? AND Name=?""",
                   (area_id, scene_name))
            if ex:
                scene_cache[cache_key] = ex[0]["PresetID"]
            else:
                so = q("""SELECT ISNULL(MAX(SortOrder),-1)+1 AS ns FROM tblPreset
                          WHERE PresetType=3 AND ParentType=2 AND ParentID=?""",
                       (area_id,))
                next_so = so[0]["ns"] if so else 0
                pid = _alloc_and_insert("tblPreset", "PresetID", {
                    "Name": scene_name, "DatabaseRevision": 0, "SortOrder": next_so,
                    "ParentID": area_id, "ParentType": 2,
                    "NeedsTransfer": 1, "PresetType": 3,
                    "WhereUsedId": 2147483647, "IsGPDPreset": 0,
                })
                scene_cache[cache_key] = pid
                created_scenes += 1

        preset_id = scene_cache[cache_key]

        # Skip rows with no zone or no level
        if not zone_name or level_str == "":
            continue

        try:
            level = int(float(level_str))
        except (ValueError, TypeError):
            errors.append(f"行{i}: レベル値が無効 '{level_str}'")
            continue
        try:
            fade = round(float(fade_str) * 4)
        except (ValueError, TypeError):
            fade = 8
        try:
            delay = round(float(delay_str) * 4)
        except (ValueError, TypeError):
            delay = 0

        # Find zone
        zr = q("""SELECT ZoneID FROM tblZone
                  WHERE ParentID=? AND Name=? AND ZoneLayer=0""",
               (area_id, zone_name))
        if not zr:
            errors.append(f"行{i}: ゾーン '{zone_name}' がエリア '{area_name}' に見つかりません")
            continue
        zone_id = zr[0]["ZoneID"]

        # Upsert assignment
        ex_assn = q("""SELECT PresetAssignmentID FROM tblPresetAssignment
                       WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=15""",
                    (preset_id, zone_id))
        try:
            if ex_assn:
                aid = ex_assn[0]["PresetAssignmentID"]
                execute_sql("UPDATE tblPresetAssignment SET NeedsTransfer=1 WHERE PresetAssignmentID=?",
                            (aid,))
                for ptype, pval in [(3, level), (1, fade), (2, delay)]:
                    ep = q("""SELECT 1 FROM tblAssignmentCommandParameter
                              WHERE ParentId=? AND ParameterType=?""", (aid, ptype))
                    if ep:
                        execute_sql("""UPDATE tblAssignmentCommandParameter
                                       SET ParameterValue=? WHERE ParentId=? AND ParameterType=?""",
                                    (pval, aid, ptype))
                    else:
                        execute_sql("""INSERT INTO tblAssignmentCommandParameter
                                       (SortOrder,ParentId,ParameterType,ParameterValue)
                                       VALUES (?,?,?,?)""", (0, aid, ptype, pval))
            else:
                sort_r = q("SELECT ISNULL(MAX(SortOrder),0)+1 AS ns FROM tblPresetAssignment WHERE ParentID=?",
                           (preset_id,))
                sort = sort_r[0]["ns"] if sort_r else 0
                aid = _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", {
                    "Name": "", "DatabaseRevision": 0, "SortOrder": sort,
                    "ParentID": preset_id, "ParentType": 43,
                    "AssignableObjectID": zone_id, "AssignableObjectType": 15,
                    "AssignmentCommandType": 2, "NeedsTransfer": 1,
                    "AssignmentCommandGroup": 1, "WhereUsedId": 2147483647,
                })
                for ptype, pval in [(3, level), (1, fade), (2, delay)]:
                    execute_sql("""INSERT INTO tblAssignmentCommandParameter
                                   (SortOrder,ParentId,ParameterType,ParameterValue)
                                   VALUES (?,?,?,?)""", (0, aid, ptype, pval))
            updated_assignments += 1
        except Exception as e:
            errors.append(f"行{i}: DB書き込みエラー: {e}")

    if updated_assignments > 0 or created_scenes > 0:
        state["dirty"] = True
    return jsonify({
        "ok": True,
        "created_scenes": created_scenes,
        "updated_assignments": updated_assignments,
        "errors": errors,
    })


@app.route("/api/area/<int:area_id>/shared-presets")
def area_shared_presets(area_id):
    """Return shared presets (PresetType=3) belonging to this area."""
    if not state["db_name"]:
        return jsonify([])
    rows = q("""
        SELECT p.PresetID, p.Name, p.SortOrder
        FROM tblPreset p
        WHERE p.PresetType = 3 AND p.ParentType = 2 AND p.ParentID = ?
        ORDER BY p.Name
    """, (area_id,))
    return jsonify(rows)


@app.route("/api/area/<int:area_id>/shared-presets", methods=["POST"])
def create_shared_preset(area_id):
    """Create a new shared preset (PresetType=3) for the given area. Max 100 per area."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "名前が空です"}), 400

    # Check 100-preset limit
    count_rows = q("""
        SELECT COUNT(*) AS cnt FROM tblPreset
        WHERE PresetType = 3 AND ParentType = 2 AND ParentID = ?
    """, (area_id,))
    count = count_rows[0]["cnt"] if count_rows else 0
    if count >= 100:
        return jsonify({"error": f"このエリアのShared Sceneは上限100個です（現在{count}個）"}), 400

    # Determine next SortOrder
    so_rows = q("""
        SELECT ISNULL(MAX(SortOrder), -1) + 1 AS next_so FROM tblPreset
        WHERE PresetType = 3 AND ParentType = 2 AND ParentID = ?
    """, (area_id,))
    next_so = so_rows[0]["next_so"] if so_rows else 0

    preset_id = _alloc_and_insert("tblPreset", "PresetID", {
        "Name": name,
        "PresetType": 3,
        "ParentType": 2,
        "ParentID": area_id,
        "SortOrder": next_so,
        "NeedsTransfer": 1,
        "DatabaseRevision": 0,
        "WhereUsedId": 2147483647,
        "IsGPDPreset": 0,
    })
    return jsonify({"preset_id": preset_id, "name": name, "sort_order": next_so}), 201


@app.route("/api/preset/<int:preset_id>/copy", methods=["POST"])
def copy_preset(preset_id):
    """Copy a shared preset and all its zone assignments."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    src = q("SELECT * FROM tblPreset WHERE PresetID=? AND PresetType=3", (preset_id,))
    if not src:
        return jsonify({"error": "プリセットが見つかりません"}), 404
    s = src[0]

    so_rows = q("SELECT ISNULL(MAX(SortOrder),-1)+1 AS ns FROM tblPreset WHERE PresetType=3 AND ParentType=2 AND ParentID=?",
                (s["ParentID"],))
    next_so = so_rows[0]["ns"] if so_rows else 0

    new_pid = _alloc_and_insert("tblPreset", "PresetID", {
        "Name": s["Name"] + " のコピー",
        "PresetType": 3, "ParentType": 2, "ParentID": s["ParentID"],
        "SortOrder": next_so, "NeedsTransfer": 1, "DatabaseRevision": 0,
        "WhereUsedId": 2147483647, "IsGPDPreset": 0,
    })

    assignments = q("SELECT * FROM tblPresetAssignment WHERE ParentID=?", (preset_id,))
    for assn in assignments:
        new_aid = _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", {
            "Name": assn.get("Name", ""),
            "DatabaseRevision": 0,
            "SortOrder": assn.get("SortOrder", 0),
            "ParentID": new_pid,
            "ParentType": assn.get("ParentType", 43),
            "AssignableObjectID": assn.get("AssignableObjectID"),
            "AssignableObjectType": assn.get("AssignableObjectType"),
            "AssignmentCommandType": assn.get("AssignmentCommandType", 2),
            "NeedsTransfer": 1,
            "AssignmentCommandGroup": assn.get("AssignmentCommandGroup", 1),
            "WhereUsedId": 2147483647,
        })
        params = q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=?",
                   (assn["PresetAssignmentID"],))
        for p in params:
            execute_sql("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (0,?,?,?)",
                        (new_aid, p["ParameterType"], p["ParameterValue"]))

    return jsonify({"preset_id": new_pid, "name": s["Name"] + " のコピー"}), 201


# ── Area Scenes ────────────────────────────────────────────────────────────────

@app.route("/api/area/<int:area_id>/scenes-full")
def area_scenes_full(area_id):
    """Return scenes list and ambient zones for an area."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    scenes = q("""
        SELECT s.SceneID, s.Name, s.Number
        FROM tblSceneController sc
        JOIN tblScene s ON s.ParentSceneControllerID = sc.SceneControllerID
        WHERE sc.ParentID = ? AND sc.ParentType = 2
        ORDER BY s.Number
    """, (area_id,))
    ambient_zones = q("""
        SELECT z.ZoneID, z.Name, z.ZoneNumber AS LoadNumber, z.ControlType
        FROM tblZone z
        WHERE z.ParentID = ? AND z.ZoneLayer = 0
        ORDER BY z.ZoneNumber
    """, (area_id,))
    return jsonify({"scenes": scenes, "ambient_zones": ambient_zones})


@app.route("/api/scene/<int:scene_id>/assignments")
def scene_assignments(scene_id):
    """Return zone assignments for an area scene, keyed by ZoneID."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    rows = q("""
        SELECT pa.PresetAssignmentID, pa.AssignableObjectID AS ZoneID,
               pa.AssignmentCommandType
        FROM tblPresetAssignment pa
        WHERE pa.ParentID = ? AND pa.AssignableObjectType = 15
    """, (scene_id,))
    result = {}
    for row in rows:
        params = q("""
            SELECT ParameterType, ParameterValue
            FROM tblAssignmentCommandParameter WHERE ParentId = ?
        """, (row["PresetAssignmentID"],))
        pm = {p["ParameterType"]: p["ParameterValue"] for p in params}
        # PT=3 absent means "Unaffected" — return None so frontend can distinguish
        level = pm.get(3)  # None if not found
        result[row["ZoneID"]] = {
            "aid": row["PresetAssignmentID"],
            "cmd_type": row["AssignmentCommandType"],
            "level": level,
            "fade":  round(pm.get(1, 0) / 4, 2),
            "delay": round(pm.get(2, 0) / 4, 2),
        }
    return jsonify(result)


@app.route("/api/scene/<int:scene_id>/zone", methods=["POST"])
def add_scene_zone(scene_id):
    """Add a zone assignment to an area scene."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    zone_id = data.get("zone_id")
    if zone_id is None:
        return jsonify({"error": "zone_id が必要です"}), 400
    # Check not already assigned
    existing = q("""
        SELECT PresetAssignmentID FROM tblPresetAssignment
        WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=15
    """, (scene_id, zone_id))
    if existing:
        return jsonify({"error": "このゾーンはすでに割り当て済みです"}), 400
    level = int(data.get("level", 100))
    fade  = round(float(data.get("fade", 2)) * 4)   # seconds → 250ms units
    delay = round(float(data.get("delay", 0)) * 4)

    # Use an existing PresetAssignment row for this scene as a column template
    # to avoid NOT NULL surprises (ParentType, WhereUsedId, etc.)
    tmpl_rows = q("""
        SELECT TOP 1 * FROM tblPresetAssignment WHERE ParentID = ?
    """, (scene_id,))
    if not tmpl_rows:
        # Fall back to any scene assignment in the DB
        tmpl_rows = q("""
            SELECT TOP 1 pa.* FROM tblPresetAssignment pa
            JOIN tblScene s ON s.SceneID = pa.ParentID
        """)

    try:
        if tmpl_rows:
            row = dict(tmpl_rows[0])
            row.pop("PresetAssignmentID", None)
        else:
            row = {"WhereUsedId": 2147483647, "IsDimmerLocalLoad": 0}
        # Explicitly set all required fields (don't rely on template values)
        row["Name"] = ""
        row["ParentID"] = scene_id
        row["ParentType"] = 41          # confirmed from DB: scene assignments use 41
        row["AssignableObjectID"] = zone_id
        row["AssignableObjectType"] = 15
        row["AssignmentCommandType"] = 2
        row["AssignmentCommandGroup"] = 1
        row["DatabaseRevision"] = 0
        row["SortOrder"] = 0
        row["NeedsTransfer"] = 1
        # Clear nullable unique fields so they don't duplicate from template
        row["Xid"] = None
        row["SmartProgrammingDefaultGUID"] = None
        for f in ("TemplateID", "TemplateUsedID", "TemplateReferenceID", "TemplateInstanceNumber"):
            row[f] = None
        aid = _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", row)
        for ptype, pval in [(3, level), (1, fade), (2, delay)]:
            execute_sql("""
                INSERT INTO tblAssignmentCommandParameter
                    (SortOrder, ParentId, ParameterType, ParameterValue)
                VALUES (?, ?, ?, ?)
            """, (0, aid, ptype, pval))
        return jsonify({"aid": aid, "level": level})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/area/<int:area_id>/scenes", methods=["POST"])
def create_area_scene(area_id):
    """Create a new area scene. Max 16 scenes per area."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "名前が空です"}), 400

    # Get SceneController for this area
    sc_rows = q("""
        SELECT SceneControllerID FROM tblSceneController
        WHERE ParentID = ? AND ParentType = 2
    """, (area_id,))
    if not sc_rows:
        return jsonify({"error": "このエリアにSceneControllerがありません"}), 400
    sc_id = sc_rows[0]["SceneControllerID"]

    # Check 16-scene limit
    count_rows = q("""
        SELECT COUNT(*) AS cnt FROM tblScene
        WHERE ParentSceneControllerID = ?
    """, (sc_id,))
    count = count_rows[0]["cnt"] if count_rows else 0
    if count >= 16:
        return jsonify({"error": f"Area Sceneは最大16個です（現在{count}個）"}), 400

    # Use an existing scene as column template to avoid NOT NULL surprises
    template_rows = q("""
        SELECT TOP 1 * FROM tblScene WHERE ParentSceneControllerID = ?
        ORDER BY Number
    """, (sc_id,))

    # Next Number and SortOrder
    num_rows = q("""
        SELECT ISNULL(MAX(Number), 0) + 1 AS next_num,
               ISNULL(MAX(SortOrder), -1) + 1 AS next_so
        FROM tblScene WHERE ParentSceneControllerID = ?
    """, (sc_id,))
    next_num = num_rows[0]["next_num"] if num_rows else 1
    next_so  = num_rows[0]["next_so"]  if num_rows else 0

    try:
        if template_rows:
            row = dict(template_rows[0])
            row.pop("SceneID", None)
        else:
            row = {
                "IsDaylightingScene": 0,
                "Icon": "",
                "IsHyperionScene": 0,
            }
        # Override/set key fields for the new scene
        row["Name"] = name
        row["Number"] = next_num
        row["SortOrder"] = next_so
        row["ParentSceneControllerID"] = sc_id
        row["DatabaseRevision"] = 0
        row["NeedsTransfer"] = 1
        # Clear template linkage so new scene is standalone
        for f in ("TemplateID", "TemplateUsedID", "TemplateReferenceID",
                  "TemplateInstanceNumber", "Xid"):
            row[f] = None

        scene_id = _alloc_and_insert("tblScene", "SceneID", row)

        # Insert corresponding tblSceneTemplate row (LD requires 1-to-1 shadow entry)
        st_row = dict(row)
        st_row["SceneID"] = scene_id
        st_row["AssignmentCommandType"] = 2   # extra column in tblSceneTemplate
        col_str = ", ".join(st_row.keys())
        ph_str  = ", ".join(["?"] * len(st_row))
        execute_sql(f"INSERT INTO tblSceneTemplate ({col_str}) VALUES ({ph_str})",
                    list(st_row.values()))

        # Pre-create Unaffected (cmd_type=1) rows for all ambient zones
        # LD requires every zone to have a tblPresetAssignment row even for Unaffected
        ambient_zones = q("""
            SELECT ZoneID FROM tblZone WHERE ParentID = ? AND ZoneLayer = 0
        """, (area_id,))
        pa_tmpl_rows = q("""
            SELECT TOP 1 * FROM tblPresetAssignment WHERE ParentType = 41
        """)
        for z in ambient_zones:
            if pa_tmpl_rows:
                pa_row = dict(pa_tmpl_rows[0])
                pa_row.pop("PresetAssignmentID", None)
            else:
                pa_row = {"WhereUsedId": 2147483647, "IsDimmerLocalLoad": 0}
            pa_row["Name"] = ""
            pa_row["ParentID"] = scene_id
            pa_row["ParentType"] = 41
            pa_row["AssignableObjectID"] = z["ZoneID"]
            pa_row["AssignableObjectType"] = 15
            pa_row["AssignmentCommandType"] = 1   # Unaffected
            pa_row["AssignmentCommandGroup"] = 1
            pa_row["DatabaseRevision"] = 0
            pa_row["SortOrder"] = 0
            pa_row["NeedsTransfer"] = 1
            pa_row["Xid"] = None
            pa_row["SmartProgrammingDefaultGUID"] = None
            for f in ("TemplateID", "TemplateUsedID", "TemplateReferenceID", "TemplateInstanceNumber"):
                pa_row[f] = None
            _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", pa_row)

        return jsonify({"scene_id": scene_id, "name": name, "number": next_num})
    except Exception as e:
        import traceback
        app.logger.error("create_area_scene error: %s\n%s", e, traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route("/api/scene/<int:scene_id>/copy", methods=["POST"])
def copy_area_scene(scene_id):
    """Copy an area scene with all its zone assignments."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    try:
        src = q("SELECT * FROM tblScene WHERE SceneID=?", (scene_id,))
        if not src:
            return jsonify({"error": "シーンが見つかりません"}), 404
        s = src[0]
        ctrl_id = s["ParentSceneControllerID"]

        # Check 16-scene limit
        cnt = q("SELECT COUNT(*) AS c FROM tblScene WHERE ParentSceneControllerID=? AND Number > 0", (ctrl_id,))
        if cnt and cnt[0]["c"] >= 16:
            return jsonify({"error": "シーン数が上限(16)に達しています"}), 400

        next_num_r = q("SELECT ISNULL(MAX(Number),0)+1 AS nn FROM tblScene WHERE ParentSceneControllerID=? AND Number > 0",
                       (ctrl_id,))
        next_num = next_num_r[0]["nn"] if next_num_r else 1
        next_so_r = q("SELECT ISNULL(MAX(SortOrder),0)+1 AS ns FROM tblScene WHERE ParentSceneControllerID=?",
                      (ctrl_id,))
        next_so = next_so_r[0]["ns"] if next_so_r else 0

        new_scene_id = _alloc_and_insert("tblScene", "SceneID", {
            "Name": s["Name"] + " のコピー",
            "Number": next_num,
            "SortOrder": next_so,
            "ParentSceneControllerID": ctrl_id,
            "IsDaylightingScene": s.get("IsDaylightingScene", 0),
            "Icon": s.get("Icon", 0),
            "IsHyperionScene": s.get("IsHyperionScene", 0),
            "DatabaseRevision": 0,
            "NeedsTransfer": 1,
            "TemplateID": None, "TemplateUsedID": None,
            "TemplateReferenceID": None, "TemplateInstanceNumber": None,
            "Xid": None,
        })

        # Mirror row in tblSceneTemplate
        execute_sql("""
            INSERT INTO tblSceneTemplate
            SELECT ?, Name, SortOrder, Number, ParentSceneControllerID, DatabaseRevision,
                   NeedsTransfer, TemplateID, TemplateUsedID, TemplateReferenceID,
                   TemplateInstanceNumber, IsDaylightingScene, Icon, IsHyperionScene, Xid, 2
            FROM tblScene WHERE SceneID = ?
        """, (new_scene_id, new_scene_id))

        # Copy all zone assignments
        assignments = q("SELECT * FROM tblPresetAssignment WHERE ParentID=? AND ParentType=41", (scene_id,))
        for assn in assignments:
            new_aid = _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", {
                "Name": assn.get("Name", ""),
                "DatabaseRevision": 0,
                "SortOrder": assn.get("SortOrder", 0),
                "ParentID": new_scene_id,
                "ParentType": 41,
                "AssignableObjectID": assn.get("AssignableObjectID"),
                "AssignableObjectType": assn.get("AssignableObjectType"),
                "AssignmentCommandType": assn.get("AssignmentCommandType", 1),
                "NeedsTransfer": 1,
                "AssignmentCommandGroup": assn.get("AssignmentCommandGroup", 1),
                "WhereUsedId": 2147483647,
                "IsDimmerLocalLoad": assn.get("IsDimmerLocalLoad", 0),
                "Xid": None, "SmartProgrammingDefaultGUID": None,
                "TemplateID": None, "TemplateUsedID": None,
                "TemplateReferenceID": None, "TemplateInstanceNumber": None,
            })
            params = q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=?",
                       (assn["PresetAssignmentID"],))
            for p in params:
                execute_sql("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (0,?,?,?)",
                            (new_aid, p["ParameterType"], p["ParameterValue"]))

        return jsonify({"scene_id": new_scene_id, "name": s["Name"] + " のコピー", "number": next_num}), 201
    except Exception as e:
        import traceback
        app.logger.error("copy_area_scene error: %s\n%s", e, traceback.format_exc())
        return jsonify({"error": str(e)}), 500


@app.route("/api/scene/<int:scene_id>/name", methods=["PUT"])
def rename_scene(scene_id):
    """Rename an area scene."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    name = (request.json or {}).get("name", "").strip()
    if not name:
        return jsonify({"error": "名前が空です"}), 400
    try:
        old = q("SELECT Name FROM tblScene WHERE SceneID=?", (scene_id,))
        old_name = old[0]["Name"] if old else ""
        execute_sql("UPDATE tblScene SET Name=? WHERE SceneID=?", (name, scene_id))
        push_undo(
            [("UPDATE tblScene SET Name=? WHERE SceneID=?", (old_name, scene_id))],
            [("UPDATE tblScene SET Name=? WHERE SceneID=?", (name, scene_id))],
            f"シーン名変更 → {name}"
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Load type master ──────────────────────────

@app.route("/api/load-types")
def load_types():
    if not state["db_name"]:
        return jsonify([])
    rows = q("SELECT LoadTypeID, Description FROM lstLoadType ORDER BY LoadTypeID")
    for r in rows:
        r["ControlType"] = load_type_to_control_type(r["LoadTypeID"])
    return jsonify(rows)


@app.route("/api/zone", methods=["POST"])
def add_zone():
    """Add a new zone (tblZone + tblSwitchLeg) to an area."""
    data = request.json
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    area_id   = int(data["area_id"])
    name      = data.get("name", "New Zone").strip() or "New Zone"
    load_type = int(data.get("load_type", 98))
    zone_layer = int(data.get("zone_layer", 0))
    control_type = load_type_to_control_type(load_type)

    # Next IDs from tblNextObjectID (authoritative counter shared across all tables)
    next_obj = q("SELECT NextObjectID FROM tblNextObjectID")[0]["NextObjectID"]
    new_zone_id = next_obj
    new_sl_id   = next_obj + 1

    next_num = q(
        "SELECT ISNULL(MAX(ZoneNumber), 0) + 1 as n FROM tblZone WHERE ParentID = ?",
        (area_id,)
    )[0]["n"]
    sort_order = next_num - 1

    # Use ZoneColorInfo from an existing zone in the same area (LD default = 18)
    zci_row = q(
        "SELECT TOP 1 ZoneColorInfo FROM tblZone WHERE ParentID = ? ORDER BY ZoneNumber",
        (area_id,)
    )
    zone_color_info = zci_row[0]["ZoneColorInfo"] if zci_row else 18

    # Get ControllerID/ControllerType from tblZonable for an existing zone in this area
    zonable_row = q("""
        SELECT TOP 1 zn.ControllerID, zn.ControllerType
        FROM tblZonable zn
        JOIN tblZone tz ON zn.AssociatedZoneID = tz.ZoneID
        WHERE tz.ParentID = ?
        ORDER BY tz.ZoneNumber
    """, (area_id,))
    controller_id   = zonable_row[0]["ControllerID"]   if zonable_row else None
    controller_type = zonable_row[0]["ControllerType"] if zonable_row else 3

    # Get LoadTypePropertyType and Voltage from existing fixture with same load type
    fixture_props = q(
        "SELECT TOP 1 LoadTypePropertyType, Voltage FROM tblFixture WHERE LoadType = ?",
        (load_type,)
    )
    ltp_type = fixture_props[0]["LoadTypePropertyType"] if fixture_props else 4
    voltage  = fixture_props[0]["Voltage"]              if fixture_props else 2

    # Get tblFixtureLighting template from existing fixture with same load type
    fl_props = q("""
        SELECT TOP 1 fl.*
        FROM tblFixtureLighting fl
        JOIN tblFixture f ON fl.FixtureID = f.FixtureID
        WHERE f.LoadType = ?
        ORDER BY f.FixtureID
    """, (load_type,))
    fl = fl_props[0] if fl_props else None

    new_fa_id      = next_obj + 2
    new_fixture_id = next_obj + 3

    # Xid: 22-char URL-safe base64
    import base64
    xid_z       = base64.urlsafe_b64encode(uuid.uuid4().bytes)[:22].decode()
    xid_sl      = base64.urlsafe_b64encode(uuid.uuid4().bytes)[:22].decode()
    xid_fa      = base64.urlsafe_b64encode(uuid.uuid4().bytes)[:22].decode()
    xid_fixture = base64.urlsafe_b64encode(uuid.uuid4().bytes)[:22].decode()

    execute_sqls([
        ("""
            INSERT INTO tblZone (
                ZoneID, ParentID, Name, DesignRevision, DatabaseRevision,
                ZoneNumber, SortOrder, ZoneDescription,
                RaiseLowerConfiguration, ControlType, ObjectType, WhereUsedId,
                ZoneColorInfo, ObjectActivationState, ZoneConfiguration,
                ZoneLayer, Guid, Xid
            ) VALUES (
                ?, ?, ?, 1, 0,
                ?, ?, '',
                0, ?, 15, 2147483647,
                ?, 0, 1,
                ?, NEWID(), ?
            )
        """, (new_zone_id, area_id, name, next_num, sort_order, control_type, zone_color_info, zone_layer, xid_z)),
        ("""
            INSERT INTO tblSwitchLeg (
                SwitchLegID, ParentID, Name, DesignRevision, DatabaseRevision,
                SortOrder, OutputNumberOnLink, AbsoluteMinimumLevel, BurnInTime,
                ElectronicBypassLevel, HighEnd, InrushDelay, LampRunHoursThreshold,
                LowEnd, ManualOverrideLevel, AbsoluteMaximumLevel, IsNightLight,
                EmergencyModeType, ProgrammedOffLevel, LoadType, ObjectType, AFCI,
                LampLifeExpectancy, LampPreWarningTime, WhereUsedId, ObjectActivationState, Xid
            ) VALUES (
                ?, ?, ?, 1, 0,
                ?, 65535, 0, 100,
                0, 100, 0, 10000,
                1, 100, 100, 0,
                1, 0, ?, 10, 0,
                20000, 100, 2147483647, 0, ?
            )
        """, (new_sl_id, area_id, str(next_num), sort_order, load_type, xid_sl)),
        # tblZonable: links zone to its controller (ZonableID = SwitchLegID, same as LD format)
        ("""
            INSERT INTO tblZonable (ZonableID, AssociatedZoneID, ControllerID, ControllerType, ZonableObjectType)
            VALUES (?, ?, ?, ?, 10)
        """, (new_sl_id, new_zone_id, controller_id, controller_type)),
        # tblFixture must be inserted BEFORE tblFixtureAssignment (FK constraint)
        # tblFixture: ParentID = FixtureAssignmentID, ParentType = 7, ObjectType = 6
        ("""
            INSERT INTO tblFixture (
                FixtureID, Name, DesignRevision, DatabaseRevision,
                ManufacturerModel, ManufacturerName, Notes,
                PriceCurrency, PriceValue, LoadTypePropertyType,
                LoadType, Voltage, FixtureWattage, SortOrder,
                ParentID, ParentType, ObjectType, FixtureDescription,
                FixtureInfoID, PhaseControl,
                AssociatedFixtureGroupId, FixtureControllerModelInfo,
                WhereUsedId, Xid
            ) VALUES (
                ?, 'Override Fixture', 1, 0,
                '', '', '',
                '', 0, ?,
                ?, ?, 0, -1,
                ?, 7, 6, '',
                0, 0,
                NULL, NULL,
                2147483647, ?
            )
        """, (new_fixture_id, ltp_type, load_type, voltage, new_fa_id, xid_fixture)),
        # tblFixtureAssignment: ParentID = SwitchLegID, ParentType = 10
        ("""
            INSERT INTO tblFixtureAssignment (
                FixtureAssignmentID, ParentID, ParentType, Name, DesignRevision, DatabaseRevision,
                NumberofFixtures, SortOrder, FixtureID, WhereUsedId, Xid
            ) VALUES (?, ?, 10, 'FixtureAssignment 001', 1, 0, 1, 0, ?, 2147483647, ?)
        """, (new_fa_id, new_sl_id, new_fixture_id, xid_fa)),
        # tblFixtureLighting: copy ballast/lighting params from existing fixture with same LoadType
        ("""
            INSERT INTO tblFixtureLighting (
                FixtureID, BallastInfoModelInfoID, BallastInterfaceModelInfoID,
                BlipTimeOffset, BlipWidth, ElectronicBypassTime,
                LampQuantity, LampWattage, LoadInterfaceModelInfoID, LoadInterfaceQuantity,
                Softstart, VoltageCompensationDisabled, VoltageCompensationAlgorithm,
                BlankingPulse, FrequencyFiltering, SoftwarePll, Slushing,
                LampType, DimmingRange, LowEnd, HighEnd, PhysicalLowEnd, PhysicalHighEnd,
                AbsoluteMinimumLevel, BallastFactor, SizeID, DefaultControlsID,
                MountingTypeID, OptionsID, LampLifeExpectancy
            ) VALUES (
                ?, ?, ?,
                ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?, ?, ?, ?,
                ?, ?, ?, ?,
                ?, ?, ?
            )
        """, (
            new_fixture_id,
            fl["BallastInfoModelInfoID"]        if fl else 1856,
            fl["BallastInterfaceModelInfoID"]   if fl else None,
            fl["BlipTimeOffset"]                if fl else 0,
            fl["BlipWidth"]                     if fl else 6,
            fl["ElectronicBypassTime"]          if fl else 0,
            fl["LampQuantity"]                  if fl else 1,
            fl["LampWattage"]                   if fl else 0,
            fl["LoadInterfaceModelInfoID"]      if fl else None,
            fl["LoadInterfaceQuantity"]         if fl else 0,
            fl["Softstart"]                     if fl else True,
            fl["VoltageCompensationDisabled"]   if fl else False,
            fl["VoltageCompensationAlgorithm"]  if fl else 0,
            fl["BlankingPulse"]                 if fl else 1,
            fl["FrequencyFiltering"]            if fl else 1,
            fl["SoftwarePll"]                   if fl else 0,
            fl["Slushing"]                      if fl else 3,
            fl["LampType"]                      if fl else 0,
            fl["DimmingRange"]                  if fl else 0,
            fl["LowEnd"]                        if fl else 1,
            fl["HighEnd"]                       if fl else 100,
            fl["PhysicalLowEnd"]                if fl else 2700,
            fl["PhysicalHighEnd"]               if fl else 6000,
            fl["AbsoluteMinimumLevel"]          if fl else 0,
            fl["BallastFactor"]                 if fl else 1,
            fl["SizeID"]                        if fl else 0,
            fl["DefaultControlsID"]             if fl else 0,
            fl["MountingTypeID"]                if fl else 0,
            fl["OptionsID"]                     if fl else 0,
            fl["LampLifeExpectancy"]            if fl else 20000,
        )),
        # tblDaylightable: DaylightableID = SwitchLegID, ObjectType=10 (switchleg)
        ("""
            INSERT INTO tblDaylightable (DaylightableID, DaylightableObjectType, GainGroupID, DaylightingDesignType)
            VALUES (?, 10, NULL, 1)
        """, (new_sl_id,)),
        # Advance the global ID counter by 4 (zone, switchleg, fixture_assignment, fixture)
        ("UPDATE tblNextObjectID SET NextObjectID = ?", (next_obj + 4,)),
    ])

    return jsonify({"ok": True, "zone_id": new_zone_id, "zone_number": next_num})


# ── Zone editing ──────────────────────────────

@app.route("/api/zone/<int:zone_id>", methods=["PATCH"])
def update_zone(zone_id):
    """Update zone name and/or description."""
    data = request.json
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    fields, params = [], []
    if "Name" in data:
        fields.append("Name = ?")
        params.append(data["Name"])
    if "ZoneDescription" in data:
        fields.append("ZoneDescription = ?")
        params.append(data["ZoneDescription"])
    if "ZoneLayer" in data:
        fields.append("ZoneLayer = ?")
        params.append(int(data["ZoneLayer"]))
    if not fields:
        return jsonify({"error": "変更なし"}), 400

    old = q(f"SELECT {', '.join(f.split(' = ')[0] for f in fields)} FROM tblZone WHERE ZoneID=?", (zone_id,))
    old_vals = list(old[0].values()) if old else [None] * len(fields)
    old_fields_sql = ", ".join(f"{f.split(' = ')[0]} = ?" for f in fields)
    params.append(zone_id)
    execute_sql(f"UPDATE tblZone SET {', '.join(fields)} WHERE ZoneID = ?", params)
    push_undo(
        [(f"UPDATE tblZone SET {old_fields_sql} WHERE ZoneID=?", old_vals + [zone_id])],
        [(f"UPDATE tblZone SET {', '.join(fields)} WHERE ZoneID=?", params)],
        "ゾーン変更"
    )
    return jsonify({"ok": True})


# ── Zone delete ───────────────────────────────

@app.route("/api/zone/<int:zone_id>", methods=["DELETE"])
def delete_zone(zone_id):
    """Delete a zone and its associated SwitchLeg."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    zones = q("SELECT ZoneID FROM tblZone WHERE ZoneID = ?", (zone_id,))
    if not zones:
        return jsonify({"error": "Zone not found"}), 404

    # SwitchLeg relationship: SwitchLegID = ZoneID + 1 (LD format)
    # Also clean up any wrongly-created SwitchLeg with ParentID=zone_id
    execute_sqls([
        ("DELETE FROM tblFixtureLighting WHERE FixtureID = ?", (zone_id + 3,)),
        ("DELETE FROM tblFixture WHERE ParentID = ? AND ParentType = 7", (zone_id + 2,)),
        ("DELETE FROM tblFixtureAssignment WHERE ParentID = ? AND ParentType = 10", (zone_id + 1,)),
        ("DELETE FROM tblDaylightable WHERE DaylightableID = ?", (zone_id + 1,)),
        ("DELETE FROM tblZonable WHERE AssociatedZoneID = ? OR ZonableID = ?",
         (zone_id, zone_id + 1)),
        ("DELETE FROM tblSwitchLeg WHERE SwitchLegID = ? OR ParentID = ?",
         (zone_id + 1, zone_id)),
        ("DELETE FROM tblZone WHERE ZoneID = ?", (zone_id,)),
    ])
    return jsonify({"ok": True})


# ── Zone debug/compare ────────────────────────

def _try_q(sql, params=()):
    try:
        return q(sql, params)
    except Exception as e:
        return [{"__error__": str(e)}]


@app.route("/api/zone/<int:zone_id>/fix-switchleg", methods=["POST"])
def fix_zone_switchleg(zone_id):
    """Fix a zone whose SwitchLeg has ParentID=zone_id instead of area_id."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    zone = q("SELECT ZoneID, ParentID, ZoneNumber FROM tblZone WHERE ZoneID = ?", (zone_id,))
    if not zone:
        return jsonify({"error": "Zone not found"}), 404
    area_id    = zone[0]["ParentID"]
    zone_number = zone[0]["ZoneNumber"]
    sl_id = zone_id + 1
    execute_sqls([
        ("UPDATE tblZone SET ObjectType = 15 WHERE ZoneID = ? AND ObjectType != 15",
         (zone_id,)),
        ("UPDATE tblSwitchLeg SET ParentID = ?, Name = ? WHERE SwitchLegID = ?",
         (area_id, str(zone_number), sl_id)),
    ])
    return jsonify({"ok": True, "fixed_sl_id": sl_id, "new_parent_id": area_id})


@app.route("/api/zone/<int:zone_id>/debug")
def debug_zone(zone_id):
    """Return full data for a zone across multiple tables (for diagnosis)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    zone = _try_q("SELECT * FROM tblZone WHERE ZoneID = ?", (zone_id,))
    sl_by_parent = _try_q("SELECT * FROM tblSwitchLeg WHERE ParentID = ?", (zone_id,))
    sl_by_id = _try_q("SELECT * FROM tblSwitchLeg WHERE SwitchLegID = ?", (zone_id + 1,))
    # Preset entries referencing this zone
    presets = _try_q("SELECT * FROM tblPreset WHERE ZoneID = ?", (zone_id,))
    # SwitchLeg by parent (zone's area)
    area_id = zone[0]["ParentID"] if zone and "ParentID" in zone[0] else None
    orphan_sls = []
    if area_id:
        orphan_sls = _try_q(
            "SELECT * FROM tblSwitchLeg WHERE ParentID = ?", (area_id,)
        )
    return jsonify({
        "zone": zone[0] if zone else None,
        "switchleg_by_parentid": sl_by_parent,
        "switchleg_by_id_plus1": sl_by_id,
        "presets_for_zone": presets,
        "switchlegs_with_area_as_parent": orphan_sls,
    })


@app.route("/api/schema-inspect")
def schema_inspect():
    """List tables that have columns likely referencing zones (ZoneID, SwitchLegID, etc.)"""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    # All user tables
    tables = _try_q("""
        SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_TYPE='BASE TABLE'
        ORDER BY TABLE_NAME
    """)

    # Columns named ZoneID or SwitchLegID anywhere
    zone_cols = _try_q("""
        SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE COLUMN_NAME IN ('ZoneID','SwitchLegID','ZoneId','SwitchlegId','OutputID')
           OR COLUMN_NAME LIKE '%ZoneID%'
           OR COLUMN_NAME LIKE '%SwitchLeg%'
        ORDER BY TABLE_NAME, COLUMN_NAME
    """)

    return jsonify({
        "all_tables": [r["TABLE_NAME"] for r in tables if "TABLE_NAME" in r],
        "zone_related_columns": zone_cols,
    })


@app.route("/api/db-integrity")
def db_integrity():
    """Check tblGuidManager, tblNextObjectID, FK violations, and triggers."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    guid_manager_cols = _try_q("""
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tblGuidManager' ORDER BY ORDINAL_POSITION
    """)
    guid_sample = _try_q("SELECT TOP 3 * FROM tblGuidManager")
    next_id = _try_q("SELECT * FROM tblNextObjectID")
    triggers = _try_q("""
        SELECT t.name, OBJECT_NAME(t.parent_id) as table_name
        FROM sys.triggers t
        WHERE OBJECT_NAME(t.parent_id) IN ('tblZone','tblSwitchLeg')
    """)
    fk_violations = _try_q("""
        SELECT fk.name as constraint_name,
               tp.name as parent_table, cp.name as parent_col,
               tr.name as ref_table, cr.name as ref_col
        FROM sys.foreign_keys fk
        JOIN sys.tables tp ON fk.parent_object_id = tp.object_id
        JOIN sys.tables tr ON fk.referenced_object_id = tr.object_id
        JOIN sys.foreign_key_columns fkc ON fk.object_id = fkc.constraint_object_id
        JOIN sys.columns cp ON fkc.parent_object_id = cp.object_id
                            AND fkc.parent_column_id = cp.column_id
        JOIN sys.columns cr ON fkc.referenced_object_id = cr.object_id
                            AND fkc.referenced_column_id = cr.column_id
        WHERE tp.name IN ('tblZone','tblSwitchLeg')
    """)
    return jsonify({
        "guid_manager_columns": [r.get("COLUMN_NAME") for r in guid_manager_cols],
        "guid_manager_sample": guid_sample,
        "next_object_id": next_id,
        "triggers_on_zone_switchleg": triggers,
        "fk_constraints": fk_violations,
    })


@app.route("/api/table-search", methods=["POST"])
def table_search():
    """Search a table for rows where a given column matches any of the provided IDs."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json
    table  = data.get("table", "")
    column = data.get("column", "")
    ids    = data.get("ids", [])
    if not table or not column or not ids:
        return jsonify({"rows": []})
    # Validate table/column names (allow only alphanumeric + underscore)
    if not re.match(r'^[A-Za-z0-9_]+$', table) or not re.match(r'^[A-Za-z0-9_]+$', column):
        return jsonify({"error": "invalid name"}), 400
    ph = ",".join("?" * len(ids))
    rows = _try_q(f"SELECT TOP 20 * FROM [{table}] WHERE [{column}] IN ({ph})", ids)
    return jsonify({"rows": rows})


@app.route("/api/table-schema/<table_name>")
def table_schema(table_name):
    """Show all column names for a given table."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    cols = _try_q("""
        SELECT COLUMN_NAME, DATA_TYPE
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = ?
        ORDER BY ORDINAL_POSITION
    """, (table_name,))
    sample = _try_q(f"SELECT TOP 3 * FROM [{table_name}]")
    return jsonify({"columns": cols, "sample": sample})


@app.route("/api/zone-cross-table/<int:zone_id>")
def zone_cross_table(zone_id):
    """Check all zone-related tables for a given zone (to find missing entries)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    sl_id = zone_id + 1
    return jsonify({
        "zone":                _try_q("SELECT * FROM tblZone WHERE ZoneID = ?", (zone_id,)),
        "switchleg":           _try_q("SELECT * FROM tblSwitchLeg WHERE SwitchLegID = ?", (sl_id,)),
        "zonable":             _try_q("SELECT * FROM tblZonable WHERE AssociatedZoneID = ?", (zone_id,)),
        "zone_control_ui":     _try_q("SELECT * FROM tblZoneControlUI WHERE AssignedZoneID = ?", (zone_id,)),
        "zone_hvac":           _try_q("SELECT * FROM tblZoneHVAC WHERE ZoneID = ?", (zone_id,)),
        "zone_phantom_hvac":   _try_q("SELECT * FROM tblZonePhantomHVAC WHERE ZoneID = ?", (zone_id,)),
        "dl_testpoint_sl":     _try_q("SELECT * FROM tblDLTestpointSwitchLeg WHERE AssignedSwitchLegID = ?", (sl_id,)),
        "shade_switchleg":     _try_q("SELECT * FROM tblShadeSwitchLeg WHERE SwitchLegID = ?", (sl_id,)),
        "sl_group_assoc":      _try_q("SELECT * FROM tblSwitchlegGroupAssociation WHERE AssociatedSwitchlegId = ?", (sl_id,)),
        # tblPreset: ParentType=15 → zone, ParentType=10 → switchleg
        "presets_zone":        _try_q("SELECT * FROM tblPreset WHERE ParentID = ? AND ParentType = 15", (zone_id,)),
        "presets_sl":          _try_q("SELECT * FROM tblPreset WHERE ParentID = ? AND ParentType = 10", (sl_id,)),
        "domain_obj":          _try_q("SELECT * FROM tblDomain WHERE DomainID = ?", (zone_id,)),
        "graphical_obj":       _try_q("SELECT * FROM tblGraphicalObject WHERE ParentID = ?", (zone_id,)),
        # tblFixtureAssignment: ParentType=10 → SwitchLeg; FixtureID = ZoneID+3
        "fixture_assignment":   _try_q("SELECT * FROM tblFixtureAssignment WHERE ParentID = ? AND ParentType = 10", (sl_id,)),
        "fixture":              _try_q("SELECT * FROM tblFixture WHERE FixtureID = ?", (zone_id + 3,)),
        # Fixture sub-tables (keyed by FixtureID = zone_id+3)
        "fixture_lighting":     _try_q("SELECT * FROM tblFixtureLighting WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_led":          _try_q("SELECT * FROM tblFixtureLed WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_led_classic":  _try_q("SELECT * FROM tblFixtureLedClassic WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_cco":          _try_q("SELECT * FROM tblFixtureCCO WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_fan":          _try_q("SELECT * FROM tblFixtureFan WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_hvac":         _try_q("SELECT * FROM tblFixtureHvac WHERE FixtureID = ?", (zone_id + 3,)),
        "lutron_fixture":       _try_q("SELECT * FROM tblLutronFixture WHERE FixtureID = ?", (zone_id + 3,)),
        "fixture_group":        _try_q("SELECT * FROM tblFixtureGroup WHERE ParentID = ? AND ParentType = 6", (zone_id + 3,)),
        "power_on_state":       _try_q("SELECT * FROM tblPowerOnState WHERE ParentID = ?", (sl_id,)),
        "daylightable":         _try_q("SELECT * FROM tblDaylightable WHERE DaylightableID = ?", (sl_id,)),
        "object_processor_map": _try_q("SELECT * FROM tblObjectToProcessorMap WHERE DomainObjectID = ?", (zone_id,)),
    })


def _build_trigger_tree(trigger_rows, preset_ids, depth=0):
    """Recursively build trigger→action tree, collecting preset IDs along the way."""
    if not trigger_rows or depth > 20:
        return []

    trig_ids = [r["TriggerID"] for r in trigger_rows]
    ph = ",".join("?" * len(trig_ids))

    # SELECT * so all condition columns (LED reference, time range, etc.) flow to the frontend
    action_rows = _try_q(f"""
        SELECT * FROM tblAction
        WHERE ParentID IN ({ph}) AND ParentType = 232
        ORDER BY ParentID, SortOrder
    """, trig_ids)

    actions_by_trig = {}
    cond_action_ids = []
    if action_rows and not action_rows[0].get("__error__"):
        for r in action_rows:
            k = r["ParentID"]
            if k not in actions_by_trig:
                actions_by_trig[k] = []
            act = dict(r)          # pass ALL columns through
            act["sub_triggers"] = []
            if r.get("PresetId"):
                preset_ids.add(r["PresetId"])
            if r["ObjectType"] == 233:
                cond_action_ids.append(r["ActionID"])
            actions_by_trig[k].append(act)

    # For conditional actions: fetch evaluations + recursively get sub-triggers
    if cond_action_ids:
        ph2 = ",".join("?" * len(cond_action_ids))

        # Fetch tblEvaluation rows for these conditional actions
        eval_rows = _try_q(f"""
            SELECT e.*,
                   v.Name    AS _var_name,
                   vs.Name   AS _state_name,
                   l.LedNumber AS _led_number,
                   cs.Name   AS _led_station_name,
                   COALESCE(z.Name, sc_cci.Name) AS _zone_name,
                   a.Name    AS _area_name,
                   og.Name   AS _occ_group_name,
                   shg.Name  AS _shade_name,
                   tc.Name   AS _tc_name,
                   seq.Name  AS _seq_name,
                   sst.Name  AS _step_name,
                   cs18.Name AS _dev_station_name
            FROM tblEvaluation e
            LEFT JOIN tblVariable            v    ON e.ConditionType = 0 AND e.FirstOperandObjectType = 169
                                                  AND e.FirstOperandObjectID = v.VariableID
            LEFT JOIN tblVariableState       vs   ON e.ConditionType = 0
                                                  AND e.SecondOperand = vs.VariableStateID
            LEFT JOIN tblLed                 l    ON e.ConditionType = 5 AND e.FirstOperandObjectType = 107
                                                  AND e.FirstOperandObjectID = l.LedID
            LEFT JOIN tblControlStationDevice csd ON e.ConditionType = 5 AND l.ParentDeviceID = csd.ControlStationDeviceID
            LEFT JOIN tblControlStation      cs   ON csd.ParentControlStationID = cs.ControlStationID
            LEFT JOIN tblZone                z    ON e.ConditionType IN (4, 7, 14) AND e.FirstOperandObjectID = z.ZoneID
            LEFT JOIN tblSensorConnection    sc_cci ON e.ConditionType = 3 AND e.FirstOperandObjectType = 66
                                                  AND e.FirstOperandObjectID = sc_cci.SensorConnectionID
            LEFT JOIN tblArea                a    ON e.ConditionType = 6 AND e.FirstOperandObjectType = 2
                                                  AND e.FirstOperandObjectID = a.AreaID
            LEFT JOIN tblOccupancyGroup      og   ON e.ConditionType = 2 AND e.FirstOperandObjectType = 38
                                                  AND e.FirstOperandObjectID = og.OccupancyGroupID
            LEFT JOIN tblShadeGroup          shg  ON e.ConditionType = 8 AND e.FirstOperandObjectType = 133
                                                  AND e.FirstOperandObjectID = shg.ShadeGroupID
            LEFT JOIN tblTimeClock           tc   ON e.ConditionType = 9 AND e.FirstOperandObjectType = 19
                                                  AND e.FirstOperandObjectID = tc.TimeClockID
            LEFT JOIN tblSequence            seq  ON e.ConditionType = 10 AND e.FirstOperandObjectType = 90
                                                  AND e.FirstOperandObjectID = seq.SequenceID
            LEFT JOIN tblSequenceStep        sst  ON e.ConditionType = 10
                                                  AND e.SecondOperand = sst.SequenceStepID
            LEFT JOIN tblControlStationDevice csd18 ON e.ConditionType = 18 AND e.FirstOperandObjectType = 5
                                                  AND e.FirstOperandObjectID = csd18.ControlStationDeviceID
            LEFT JOIN tblControlStation      cs18 ON csd18.ParentControlStationID = cs18.ControlStationID
            WHERE e.ParentID IN ({ph2}) AND e.ParentType = 233
            ORDER BY e.ParentID, e.SortOrder
        """, cond_action_ids)

        evals_by_action = {}
        if eval_rows and not eval_rows[0].get("__error__"):
            for r in eval_rows:
                k = r["ParentID"]
                if k not in evals_by_action:
                    evals_by_action[k] = []
                evals_by_action[k].append(dict(r))

        # Attach evaluations to each conditional action
        for acts in actions_by_trig.values():
            for act in acts:
                if act["ObjectType"] == 233:
                    act["evaluations"] = evals_by_action.get(act["ActionID"], [])

        sub_trig_rows = _try_q(f"""
            SELECT * FROM tblTrigger
            WHERE ParentId IN ({ph2})
            ORDER BY ParentId, TriggerType, SortOrder
        """, cond_action_ids)

        if sub_trig_rows and not sub_trig_rows[0].get("__error__"):
            by_cond = {}
            for r in sub_trig_rows:
                k = r["ParentId"]
                if k not in by_cond:
                    by_cond[k] = []
                trig_entry = dict(r)
                trig_entry["actions"] = []
                by_cond[k].append(trig_entry)

            # Recursively build sub-trees and attach
            for cond_id, sub_trigs in by_cond.items():
                sub_tree = _build_trigger_tree(sub_trigs, preset_ids, depth + 1)
                for acts in actions_by_trig.values():
                    for act in acts:
                        if act["ActionID"] == cond_id:
                            act["sub_triggers"] = sub_tree

    # Build result
    result = []
    for trig in trigger_rows:
        trig_id = trig["TriggerID"]
        result.append({
            "TriggerID": trig_id,
            "TriggerType": trig["TriggerType"],
            "SortOrder": trig.get("SortOrder", 0),
            "actions": actions_by_trig.get(trig_id, []),
        })
    return result


@app.route("/api/conditional-detail/<int:action_id>")
def conditional_detail(action_id):
    """Return all columns of a conditional action and its sub-triggers for condition text research."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    act = _try_q("SELECT * FROM tblAction WHERE ActionID = ?", (action_id,))
    # Get all columns of tblAction
    cols = _try_q("""
        SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tblAction' ORDER BY ORDINAL_POSITION
    """)
    sub_trigs = _try_q("""
        SELECT * FROM tblTrigger WHERE ParentId = ? ORDER BY TriggerType
    """, (action_id,))
    return jsonify({
        "action_columns": [c["COLUMN_NAME"] for c in (cols or [])],
        "action": act[0] if act else None,
        "sub_triggers": sub_trigs,
    })


@app.route("/api/programming-schema")
def programming_schema():
    """Explore all programming-related tables in the database."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    prog_tables = _try_q("""
        SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_TYPE='BASE TABLE'
        AND (TABLE_NAME LIKE '%Program%'
             OR TABLE_NAME LIKE '%Action%'
             OR TABLE_NAME LIKE '%Event%'
             OR TABLE_NAME LIKE '%Condition%'
             OR TABLE_NAME LIKE '%Button%'
             OR TABLE_NAME LIKE '%Trigger%'
             OR TABLE_NAME LIKE '%Step%'
             OR TABLE_NAME LIKE '%Sequence%')
        ORDER BY TABLE_NAME
    """)

    result = {}
    for tbl_row in prog_tables:
        name = tbl_row["TABLE_NAME"]
        cols = _try_q(f"""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = N'{name}'
            ORDER BY ORDINAL_POSITION
        """)
        sample = _try_q(f"SELECT TOP 5 * FROM [{name}]")
        result[name] = {
            "columns": [c["COLUMN_NAME"] for c in cols],
            "sample": sample,
        }

    return jsonify(result)


@app.route("/api/button/<int:button_id>/program")
def button_program(button_id):
    """Get programming details for a specific button."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    button = _try_q("""
        SELECT kb.ButtonID, kb.ButtonNumber, kb.Name as ButtonName,
               kb.ParentDeviceID, kb.ProgrammingModelID,
               pm.Name as ProgModel, pm.ControlType as ProgControlType, pm.PresetID
        FROM tblKeypadButton kb
        LEFT JOIN tblProgrammingModel pm ON kb.ProgrammingModelID = pm.ProgrammingModelID
        WHERE kb.ButtonID = ?
    """, (button_id,))

    pm_id = button[0]["ProgrammingModelID"] if button else None

    # Try to get events/steps from various potential table names
    events = []
    action_steps = []
    conditions = []

    if pm_id:
        # Try common Lutron programming table patterns
        for tbl in ["tblKeypadButtonEvent", "tblButtonEvent", "tblProgrammingModelEvent"]:
            rows = _try_q(f"""
                SELECT t.* FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_NAME = N'{tbl}' AND TABLE_TYPE='BASE TABLE'
            """)
            if rows and not rows[0].get("__error__"):
                events = _try_q(f"SELECT * FROM [{tbl}] WHERE ProgrammingModelID = ?", (pm_id,))
                if events and not events[0].get("__error__"):
                    break

        for tbl in ["tblActionStep", "tblProgramStep", "tblActionGroup", "tblProgrammingStep"]:
            rows = _try_q(f"""
                SELECT t.* FROM INFORMATION_SCHEMA.TABLES t
                WHERE TABLE_NAME = N'{tbl}' AND TABLE_TYPE='BASE TABLE'
            """)
            if rows and not rows[0].get("__error__"):
                action_steps = _try_q(f"SELECT * FROM [{tbl}] WHERE ProgrammingModelID = ?", (pm_id,))
                if action_steps and not action_steps[0].get("__error__"):
                    break

        for tbl in ["tblConditional", "tblConditionGroup", "tblProgrammingCondition"]:
            rows = _try_q(f"""
                SELECT t.* FROM INFORMATION_SCHEMA.TABLES t
                WHERE TABLE_NAME = N'{tbl}' AND TABLE_TYPE='BASE TABLE'
            """)
            if rows and not rows[0].get("__error__"):
                conditions = _try_q(f"SELECT * FROM [{tbl}] WHERE ProgrammingModelID = ?", (pm_id,))
                if conditions and not conditions[0].get("__error__"):
                    break

    return jsonify({
        "button": button[0] if button else None,
        "events": events,
        "action_steps": action_steps,
        "conditions": conditions,
    })


@app.route("/api/area/<int:area_id>/programs")
def area_programs(area_id):
    """Get all button programming for keypads in an area."""
    if not state["db_name"]:
        return jsonify({})

    stations = q("""
        SELECT cs.ControlStationID, cs.Name as StationName,
               csd.ControlStationDeviceID, csd.Name as DeviceName, csd.ModelInfoID
        FROM tblControlStation cs
        LEFT JOIN tblControlStationDevice csd ON csd.ParentControlStationID = cs.ControlStationID
        WHERE cs.ParentId = ? AND cs.ParentType = 2
        ORDER BY cs.SortOrder, cs.Name
    """, (area_id,))

    station_ids = list({s["ControlStationDeviceID"] for s in stations if s["ControlStationDeviceID"]})
    if not station_ids:
        return jsonify({"stations": stations, "buttons": [], "pm_details": {}, "preset_names": {}})

    ph = ",".join("?" * len(station_ids))
    buttons = q(f"""
        SELECT kb.ButtonID, kb.ButtonNumber, kb.Name as ButtonName,
               kb.ParentDeviceID, kb.ProgrammingModelID,
               pm.Name as ProgModel, pm.ControlType as ProgControlType
        FROM tblKeypadButton kb
        LEFT JOIN tblProgrammingModel pm ON kb.ProgrammingModelID = pm.ProgrammingModelID
        WHERE kb.ParentDeviceID IN ({ph})
        ORDER BY kb.ParentDeviceID, kb.ButtonNumber
    """, station_ids)

    pm_ids = list({b["ProgrammingModelID"] for b in buttons if b["ProgrammingModelID"]})

    # Fetch full ProgrammingModel details for all buttons
    pm_details = {}
    if pm_ids:
        ph2 = ",".join("?" * len(pm_ids))
        pm_rows = _try_q(f"""
            SELECT ProgrammingModelID, Name, LedLogic, UseReverseLedLogic,
                   PresetID, PressPresetID, ReleasePresetID, HoldPresetId,
                   DoubleTapPresetID, OnPresetID, OffPresetID,
                   ReferencePresetIDForLed, AllowDoubleTap, HoldTime,
                   HeldButtonAction, ControlType, VariableId, Direction,
                   ThreeWayToggle, ParentID, ParentType
            FROM tblProgrammingModel
            WHERE ProgrammingModelID IN ({ph2})
        """, pm_ids)
        if pm_rows and not pm_rows[0].get("__error__"):
            for r in pm_rows:
                pm_details[r["ProgrammingModelID"]] = r

    # Collect all referenced preset IDs to look up names
    preset_id_set = set()
    for pm in pm_details.values():
        for field in ["PresetID", "PressPresetID", "ReleasePresetID", "HoldPresetId",
                      "DoubleTapPresetID", "OnPresetID", "OffPresetID", "ReferencePresetIDForLed"]:
            if pm.get(field):
                preset_id_set.add(pm[field])

    # Look up preset names from tblPreset (tblAction.PresetId always references tblPreset)
    preset_names = {}
    if preset_id_set:
        ph3 = ",".join("?" * len(preset_id_set))
        preset_list = list(preset_id_set)
        preset_rows = _try_q(f"""
            SELECT p.PresetID, p.Name, p.ParentID, p.ParentType,
                   a.Name AS AreaName
            FROM tblPreset p
            LEFT JOIN tblArea a ON p.ParentType = 2 AND p.ParentID = a.AreaID
            WHERE p.PresetID IN ({ph3})
        """, preset_list)
        if preset_rows and not preset_rows[0].get("__error__"):
            for r in preset_rows:
                preset_names[r["PresetID"]] = {
                    "name": r.get("Name") or str(r["PresetID"]),
                    "number": None,
                    "controller": r.get("AreaName"),
                }

    # Build trigger trees for all PM parents (ButtonGroup IDs and direct PM IDs)
    # Try both: ButtonGroup parent IDs and direct PM IDs as trigger parents
    parent_ids_to_try = set()
    for pm in pm_details.values():
        if pm.get("ParentID"):
            parent_ids_to_try.add(pm["ParentID"])  # ButtonGroup ID
    parent_ids_to_try.update(pm_ids)  # Also try PM IDs directly

    trigger_trees = {}  # parent_id -> [trigger tree]
    if parent_ids_to_try:
        ph_p = ",".join("?" * len(parent_ids_to_try))
        root_trig_rows = _try_q(f"""
            SELECT * FROM tblTrigger
            WHERE ParentId IN ({ph_p})
            ORDER BY ParentId, SortOrder
        """, list(parent_ids_to_try))

        if root_trig_rows and not root_trig_rows[0].get("__error__"):
            by_parent = {}
            for r in root_trig_rows:
                k = r["ParentId"]
                if k not in by_parent:
                    by_parent[k] = []
                trig_entry = dict(r)
                trig_entry["actions"] = []
                by_parent[k].append(trig_entry)

            for parent_id, trigs in by_parent.items():
                tree = _build_trigger_tree(trigs, preset_id_set)
                trigger_trees[parent_id] = tree

    # Fetch any remaining preset names collected during tree build
    if preset_id_set:
        missing_presets = [pid for pid in preset_id_set if pid not in preset_names]
        if missing_presets:
            ph_m = ",".join("?" * len(missing_presets))
            pr2 = _try_q(f"""
                SELECT p.PresetID, p.Name, p.ParentType,
                       a.Name AS AreaName
                FROM tblPreset p
                LEFT JOIN tblArea a ON p.ParentType = 2 AND p.ParentID = a.AreaID
                WHERE p.PresetID IN ({ph_m})
            """, missing_presets)
            if pr2 and not pr2[0].get("__error__"):
                for r in pr2:
                    if r["PresetID"] not in preset_names:
                        preset_names[r["PresetID"]] = {
                            "name": r.get("Name") or str(r["PresetID"]),
                            "number": None,
                            "controller": r.get("AreaName"),
                        }

    return jsonify({
        "stations": stations,
        "buttons": buttons,
        "pm_details": pm_details,
        "preset_names": preset_names,
        "trigger_trees": trigger_trees,
    })


@app.route("/api/zone-full-compare/<int:existing_id>/<int:new_id>")
def zone_full_compare(existing_id, new_id):
    """Full column-by-column comparison between two zones and their SwitchLegs."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    z1  = _try_q("SELECT * FROM tblZone WHERE ZoneID = ?", (existing_id,))
    z2  = _try_q("SELECT * FROM tblZone WHERE ZoneID = ?", (new_id,))
    sl1 = _try_q("SELECT * FROM tblSwitchLeg WHERE SwitchLegID = ?", (existing_id + 1,))
    sl2 = _try_q("SELECT * FROM tblSwitchLeg WHERE SwitchLegID = ?", (new_id + 1,))
    return jsonify({
        "existing_zone": z1[0] if z1 else None,
        "new_zone":      z2[0] if z2 else None,
        "existing_switchleg": sl1[0] if sl1 else None,
        "new_switchleg":      sl2[0] if sl2 else None,
    })


@app.route("/api/scene-template-debug")
def scene_template_debug():
    """Inspect tblSceneTemplate structure and tblPresetAssignment schema for scenes."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    # tblPresetAssignment columns + nullable info
    pa_cols = _try_q("""
        SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tblPresetAssignment'
        ORDER BY ORDINAL_POSITION
    """)

    # Existing scene assignments: what ParentType do they use?
    pa_scene_sample = _try_q("""
        SELECT TOP 5 pa.*
        FROM tblPresetAssignment pa
        JOIN tblScene s ON s.SceneID = pa.ParentID
    """)

    # tblSceneTemplate columns
    st_cols = _try_q("""
        SELECT COLUMN_NAME, DATA_TYPE, IS_NULLABLE, COLUMN_DEFAULT
        FROM INFORMATION_SCHEMA.COLUMNS
        WHERE TABLE_NAME = 'tblSceneTemplate'
        ORDER BY ORDINAL_POSITION
    """)

    # All distinct AssignmentCommandType values used in scene assignments
    pa_cmd_types = _try_q("""
        SELECT DISTINCT pa.AssignmentCommandType,
               COUNT(*) AS cnt,
               MIN(pa.PresetAssignmentID) AS sample_aid
        FROM tblPresetAssignment pa
        JOIN tblScene s ON s.SceneID = pa.ParentID
        GROUP BY pa.AssignmentCommandType
    """)

    # Scene assignments for zones — show full detail including parameter values
    # Find a scene that has the most zone assignments (likely LD-created with all zones set)
    pa_full_sample = _try_q("""
        SELECT TOP 20 pa.PresetAssignmentID, pa.ParentID AS SceneID,
               pa.AssignableObjectID AS ZoneID, pa.AssignableObjectType,
               pa.AssignmentCommandType, pa.AssignmentCommandGroup,
               acp.ParameterType, acp.ParameterValue,
               s.Name AS SceneName
        FROM tblPresetAssignment pa
        JOIN tblScene s ON s.SceneID = pa.ParentID
        LEFT JOIN tblAssignmentCommandParameter acp ON acp.ParentId = pa.PresetAssignmentID
        ORDER BY pa.ParentID, pa.PresetAssignmentID
    """)

    return jsonify({
        "pa_cmd_types_used_in_scenes": pa_cmd_types,
        "pa_full_sample": pa_full_sample,
    })


@app.route("/api/area/<int:area_id>/debug")
def debug_area(area_id):
    """Full dump of all zones+switchlegs in an area for comparison."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    zones = _try_q("SELECT * FROM tblZone WHERE ParentID = ? ORDER BY ZoneNumber", (area_id,))
    zone_ids = [z["ZoneID"] for z in zones if "ZoneID" in z]
    result = []
    for z in zones:
        zid = z.get("ZoneID")
        sl_parent = _try_q("SELECT * FROM tblSwitchLeg WHERE ParentID = ?", (zid,))
        sl_id1    = _try_q("SELECT * FROM tblSwitchLeg WHERE SwitchLegID = ?", (zid + 1,))
        presets   = _try_q("SELECT * FROM tblPreset WHERE ZoneID = ?", (zid,))
        result.append({
            "zone": z,
            "switchleg_by_parentid": sl_parent,
            "switchleg_by_id_plus1": sl_id1,
            "presets": presets,
        })
    # Orphaned SwitchLegs (ParentID = area, not a zone)
    orphan_sls = _try_q(
        "SELECT * FROM tblSwitchLeg WHERE ParentID = ?", (area_id,)
    )
    return jsonify({"zones": result, "orphaned_switchlegs": orphan_sls})


# ── Program structure helpers ─────────────────

def _next_sort_order_action(trigger_id):
    rows = _try_q("SELECT COALESCE(MAX(SortOrder)+1, 0) AS nxt FROM tblAction WHERE ParentID = ?", (trigger_id,))
    if rows and not rows[0].get("__error__") and rows[0]["nxt"] is not None:
        return rows[0]["nxt"]
    return 0


def _delete_action_recursive(action_id):
    """Delete an action + its evaluations + its sub-triggers (and their actions) recursively."""
    sub_trigs = _try_q("SELECT TriggerID FROM tblTrigger WHERE ParentId = ?", (action_id,))
    if sub_trigs and not sub_trigs[0].get("__error__"):
        for t in sub_trigs:
            _delete_trigger_recursive(t["TriggerID"])
    execute_sql("DELETE FROM tblEvaluation WHERE ParentID = ? AND ParentType = 233", (action_id,))
    execute_sql("DELETE FROM tblAction WHERE ActionID = ?", (action_id,))


def _delete_trigger_recursive(trigger_id):
    """Delete a trigger + all its actions recursively."""
    acts = _try_q("SELECT ActionID FROM tblAction WHERE ParentID = ? AND ParentType = 232", (trigger_id,))
    if acts and not acts[0].get("__error__"):
        for a in acts:
            _delete_action_recursive(a["ActionID"])
    execute_sql("DELETE FROM tblTrigger WHERE TriggerID = ?", (trigger_id,))


@app.route("/api/trigger/<int:trigger_id>/add-action", methods=["POST"])
def add_action_to_trigger(trigger_id):
    """Add a Run / Delay / If action to an existing trigger."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    action_type = data.get("type")
    sort = _next_sort_order_action(trigger_id)

    # Optional: insert before or after a specific action
    insert_after_id = data.get("insert_after_action_id")
    insert_before_id = data.get("insert_before_action_id")
    if insert_after_id is not None or insert_before_id is not None:
        ref_id = insert_after_id if insert_after_id is not None else insert_before_id
        ref_rows = _try_q("SELECT SortOrder FROM tblAction WHERE ActionID = ?", (ref_id,))
        if ref_rows and not ref_rows[0].get("__error__") and ref_rows[0].get("SortOrder") is not None:
            ref_sort = ref_rows[0]["SortOrder"]
            insert_sort = ref_sort + 1 if insert_after_id is not None else ref_sort
            execute_sql(
                "UPDATE tblAction SET SortOrder = SortOrder + 1 "
                "WHERE ParentID = ? AND ParentType = 232 AND SortOrder >= ?",
                (trigger_id, insert_sort)
            )
            sort = insert_sort

    try:
        if action_type == "run":
            preset_id = data.get("preset_id") or 0
            action_id = _alloc_and_insert("tblAction", "ActionID", {
                "ObjectType": 234, "DatabaseRevision": 0,
                "ParentID": trigger_id, "ParentType": 232, "SortOrder": sort,
                "DelayTime": 0, "ExecutionType": 1, "PresetId": preset_id,
                "WhereUsedId": 2147483647,
            })
            return jsonify({"ok": True, "action_id": action_id})

        elif action_type == "delay":
            delay_ms = max(1000, min(int(data.get("delay_ms", 1000)), 14400000))
            action_id = _alloc_and_insert("tblAction", "ActionID", {
                "ObjectType": 235, "DatabaseRevision": 0,
                "ParentID": trigger_id, "ParentType": 232, "SortOrder": sort,
                "DelayTime": delay_ms // 1000, "ExecutionType": 0, "PresetId": 0,
                "WhereUsedId": 2147483647,
            })
            return jsonify({"ok": True, "action_id": action_id})

        elif action_type == "if":
            # Conditional action
            action_id = _alloc_and_insert("tblAction", "ActionID", {
                "ObjectType": 233, "DatabaseRevision": 0,
                "ParentID": trigger_id, "ParentType": 232, "SortOrder": sort,
                "DelayTime": 0, "ExecutionType": 0, "PresetId": 0,
                "WhereUsedId": 2147483647,
            })
            # Then branch trigger (no default evaluation — user selects condition via picker)
            _alloc_and_insert("tblTrigger", "TriggerID", {
                "ObjectType": 232, "ParentId": action_id, "ParentType": 233,
                "DatabaseRevision": 0, "SortOrder": 0, "TriggerType": 5,
                "WhereUsedId": 2147483647,
            })
            return jsonify({"ok": True, "action_id": action_id})

        else:
            return jsonify({"error": f"不明なtype: {action_type}"}), 400

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/action/<int:action_id>/add-condition", methods=["POST"])
def add_condition_to_action(action_id):
    """Add an AND/OR condition (ObjectType=236 connector + ObjectType=237 condition) to an If action."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    logical_op = int(data.get("logical_operator", 1))  # 1=AND, 2=OR
    try:
        rows = _try_q(
            "SELECT COALESCE(MAX(SortOrder)+1, 1) AS nxt FROM tblEvaluation WHERE ParentID = ? AND ParentType = 233",
            (action_id,)
        )
        next_sort = rows[0]["nxt"] if rows and not rows[0].get("__error__") else 1
        # Logical connector row (ObjectType=236)
        conn_id = _alloc_and_insert("tblEvaluation", "EvaluationID", {
            "ObjectType": 236, "DatabaseRevision": 0,
            "ParentID": action_id, "ParentType": 233, "SortOrder": next_sort,
            "EvaluationOperator": logical_op,
            "FirstOperandObjectID": 0, "FirstOperandObjectType": 0,
            "FirstOperandRefProperty": 0, "SecondOperand": 0,
            "ConditionType": 0, "ThirdOperand": 0, "WhereUsedId": 2147483647,
        })
        # New condition row (ObjectType=237) — default: DND Mode = 1
        eval_id = _alloc_and_insert("tblEvaluation", "EvaluationID", {
            "ObjectType": 237, "DatabaseRevision": 0,
            "ParentID": action_id, "ParentType": 233, "SortOrder": next_sort + 1,
            "EvaluationOperator": 3, "FirstOperandObjectID": 34,
            "FirstOperandObjectType": 400, "FirstOperandRefProperty": 151,
            "SecondOperand": 1, "ConditionType": 23, "ThirdOperand": 0,
            "WhereUsedId": 2147483647,
        })
        return jsonify({"ok": True, "connector_id": conn_id, "eval_id": eval_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/action/<int:action_id>/create-evaluation", methods=["POST"])
def create_evaluation_for_action(action_id):
    """Create the first condition row for an If action.
    Time conditions use ObjectType=241, all others use ObjectType=237."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    cond_type = int(data.get("ConditionType", 23))
    # Time conditions (ConditionType=1) use ObjectType=241; others use 237
    obj_type = int(data.get("ObjectType", 241 if cond_type == 1 else 237))
    fobj_type = int(data.get("FirstOperandObjectType", 0))
    fobj_id = int(data.get("FirstOperandObjectID") or 0)
    fobj_ref = int(data.get("FirstOperandRefProperty") or 0)
    op = int(data.get("EvaluationOperator", 9 if cond_type == 1 else 3))
    second = int(data.get("SecondOperand", 0))
    third = int(data.get("ThirdOperand") or 0)
    try:
        eval_id = _alloc_and_insert("tblEvaluation", "EvaluationID", {
            "ObjectType": obj_type, "DatabaseRevision": 0,
            "ParentID": action_id, "ParentType": 233, "SortOrder": 0,
            "EvaluationOperator": op, "FirstOperandObjectID": fobj_id,
            "FirstOperandObjectType": fobj_type, "FirstOperandRefProperty": fobj_ref,
            "SecondOperand": second, "ConditionType": cond_type, "ThirdOperand": third,
            "WhereUsedId": 2147483647,
        })
        return jsonify({"ok": True, "eval_id": eval_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/trigger/<int:trigger_id>/reorder", methods=["PUT"])
def reorder_trigger_actions(trigger_id):
    """Reorder actions in a trigger by assigning new SortOrder values."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    action_ids = data.get("action_ids_ordered", [])
    if not action_ids:
        return jsonify({"error": "action_ids_ordered 必須"}), 400
    try:
        stmts = [
            ("UPDATE tblAction SET SortOrder = ? WHERE ActionID = ? AND ParentID = ? AND ParentType = 232",
             (i, aid, trigger_id))
            for i, aid in enumerate(action_ids)
        ]
        execute_sqls(stmts)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/action/<int:action_id>/add-else", methods=["POST"])
def add_else_branch(action_id):
    """Add an Else branch (TriggerType=6) to an existing conditional action."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    try:
        _alloc_and_insert("tblTrigger", "TriggerID", {
            "ObjectType": 232, "ParentId": action_id, "ParentType": 233,
            "DatabaseRevision": 0, "SortOrder": 1, "TriggerType": 6,
            "WhereUsedId": 2147483647,
        })
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/action/<int:action_id>", methods=["DELETE"])
def delete_action_endpoint(action_id):
    """Delete an action and all its children recursively."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    try:
        _delete_action_recursive(action_id)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/pm/<int:pm_id>/add-trigger", methods=["POST"])
def add_root_trigger(pm_id):
    """Add a root trigger to a PM (for event types that have no trigger yet)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    trigger_type = int(data.get("trigger_type", 1))
    try:
        rows = _try_q(
            "SELECT COALESCE(MAX(SortOrder)+1,0) AS nxt FROM tblTrigger WHERE ParentId = ?", (pm_id,))
        sort = rows[0]["nxt"] if rows and not rows[0].get("__error__") else 0
        trigger_id = _alloc_and_insert("tblTrigger", "TriggerID", {
            "ObjectType": 232, "ParentId": pm_id, "ParentType": 231,
            "DatabaseRevision": 0, "SortOrder": sort, "TriggerType": trigger_type,
            "WhereUsedId": 2147483647,
        })
        return jsonify({"ok": True, "trigger_id": trigger_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Program editing ───────────────────────────

@app.route("/api/variables")
def all_variables():
    """Return all variables with their states for the condition editor."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    rows = _try_q("""
        SELECT v.VariableID, v.Name AS VariableName,
               vs.VariableStateID, vs.Name AS StateName, vs.SortOrder
        FROM tblVariable v
        JOIN tblVariableState vs ON vs.ParentID = v.VariableID
        ORDER BY v.Name, vs.SortOrder
    """)
    variables = {}
    for r in (rows or []):
        if r.get("__error__"):
            break
        vid = r["VariableID"]
        if vid not in variables:
            variables[vid] = {"VariableID": vid, "Name": r["VariableName"], "states": []}
        variables[vid]["states"].append({
            "VariableStateID": r["VariableStateID"],
            "Name": r["StateName"],
        })
    return jsonify({"variables": list(variables.values())})


@app.route("/api/leds")
def all_leds():
    """Return all LEDs with station name for the condition editor."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    rows = _try_q("""
        SELECT l.LedID, l.LedNumber, l.ParentDeviceID,
               csd.Name AS DeviceName,
               cs.Name  AS StationName,
               a.Name   AS AreaName,
               pa.Name  AS ParentAreaName
        FROM tblLed l
        JOIN tblControlStationDevice csd ON csd.ControlStationDeviceID = l.ParentDeviceID
        JOIN tblControlStation       cs  ON cs.ControlStationID = csd.ParentControlStationID
        LEFT JOIN tblArea             a  ON cs.ParentID = a.AreaID
        LEFT JOIN tblArea            pa  ON a.ParentID = pa.AreaID
        ORDER BY pa.Name, a.Name, cs.Name, l.LedNumber
    """)
    if not rows or rows[0].get("__error__"):
        rows = _try_q("""
            SELECT LedID, LedNumber, ParentDeviceID,
                   CAST(ParentDeviceID AS nvarchar(20)) AS DeviceName,
                   CAST(ParentDeviceID AS nvarchar(20)) AS StationName,
                   NULL AS AreaName, NULL AS ParentAreaName
            FROM tblLed ORDER BY ParentDeviceID, LedNumber
        """)
    return jsonify({"leds": rows or []})


@app.route("/api/cond-data")
def cond_data():
    """Return all data needed for the condition picker (all categories)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    errors = {}
    def q(sql, params=None):
        rows = _try_q(sql, params or ())
        if not rows or rows[0].get("__error__"):
            err = (rows[0].get("__error__") if rows else "empty")
            errors[sql[:60].strip()] = err
            return []
        return [dict(r) for r in rows]

    # Occupancy groups — linked to area via PIR sensors
    occupancy = q("""
        SELECT DISTINCT og.OccupancyGroupID, og.Name AS OGName,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblOccupancyGroup og
        CROSS JOIN (
            SELECT DISTINCT cs.ParentID AS AreaID
            FROM tblSensor s
            JOIN tblControlStationDevice csd ON s.ParentID = csd.ControlStationDeviceID
            JOIN tblControlStation cs ON csd.ParentControlStationID = cs.ControlStationID
            WHERE s.ObjectType = 325
        ) sa
        JOIN tblArea a ON sa.AreaID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        ORDER BY pa.Name, a.Name
    """)

    # CCI — contact inputs: tblSensorConnection (ObjectType=66) via tblEnclosureDevice
    cci = q("""
        SELECT sc.SensorConnectionID, sc.Name AS InputName,
               e.EnclosureID, e.Name AS EnclosureName,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblSensorConnection sc
        JOIN tblEnclosureDevice ed ON sc.ParentDeviceID = ed.EnclosureDeviceID
        JOIN tblEnclosure e ON ed.ParentEnclosureID = e.EnclosureID
        LEFT JOIN tblArea a  ON e.ParentAreaID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        WHERE sc.ObjectType = 66
        ORDER BY pa.Name, a.Name, e.Name, sc.SortOrder
    """)

    # CCO zones — zones with SwitchLeg LoadType 25 (Maintained) or 26 (Pulsed)
    cco = q("""
        SELECT z.ZoneID, z.Name AS ZoneName, z.ControlType, sl.LoadType,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblZone z
        JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
        JOIN tblArea a  ON z.ParentID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        WHERE sl.LoadType IN (25, 26)
        ORDER BY pa.Name, a.Name, z.Name
    """)

    # Lighting areas — leaf-level areas (no children)
    areas = q("""
        SELECT a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblArea a
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        WHERE NOT EXISTS (SELECT 1 FROM tblArea c WHERE c.ParentID = a.AreaID)
        ORDER BY pa.Name, a.Name
    """)

    # Lighting zones — ObjectType=15 zones excluding CCO (LoadType 25/26)
    zones = q("""
        SELECT z.ZoneID, z.Name AS ZoneName, z.ControlType,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblZone z
        LEFT JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
        JOIN tblArea a  ON z.ParentID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        WHERE z.ObjectType = 15
          AND (sl.LoadType IS NULL OR sl.LoadType NOT IN (25, 26))
        ORDER BY pa.Name, a.Name, z.Name
    """)

    # Rentable spaces (hotel room properties)
    rooms = q("""
        SELECT rs.RentableSpaceID, rs.Name AS RoomName
        FROM tblRentableSpace rs
        ORDER BY rs.Name
    """)

    # Integration command sets (Ethernet Devices, ObjType=202)
    integrations = q("""
        SELECT ics.IntegrationCommandSetID, ics.Name AS CommandName,
               ip.Name AS PortName,
               a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblIntegrationCommandSet ics
        JOIN tblIntegrationPort ip ON ics.ParentIntegrationPortID = ip.IntegrationPortID
        LEFT JOIN tblIntegrationPortAreaAssn ipaa ON ipaa.IntegrationPortID = ip.IntegrationPortID
        LEFT JOIN tblArea a  ON ipaa.AreaID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        ORDER BY ip.Name, ics.Name
    """)

    # Shade groups
    shades = q("""
        SELECT sg.ShadeGroupID, sg.Name AS ShadeName,
               a.Name AS AreaName
        FROM tblShadeGroup sg
        LEFT JOIN tblArea a ON sg.ParentAreaID = a.AreaID
        ORDER BY a.Name, sg.Name
    """)

    # Sequences
    sequences = q("""
        SELECT s.SequenceID, s.Name AS SequenceName,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblSequence s
        LEFT JOIN tblArea a  ON s.ParentID = a.AreaID AND s.ParentType = 2
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        ORDER BY s.Name
    """)

    # Timeclocks with modes
    timeclocks = q("""
        SELECT tc.TimeClockID, tc.Name AS TimeclockName,
               tm.TimeClockModeID, tm.Name AS ModeName, tm.SortOrder AS ModeSort
        FROM tblTimeClock tc
        LEFT JOIN tblTimeClockMode tm ON tm.ParentTimeclockID = tc.TimeClockID
        ORDER BY tc.Name, tm.SortOrder
    """)

    # HVAC zones — via tblZoneHVAC (more reliable than ObjectType=211 filter)
    hvac = q("""
        SELECT z.ZoneID, z.Name AS ZoneName,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblZoneHVAC hz
        JOIN tblZone z ON hz.ZoneID = z.ZoneID
        JOIN tblArea a  ON z.ParentID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        ORDER BY pa.Name, a.Name, z.Name
    """)
    if not hvac:
        hvac = q("""
            SELECT z.ZoneID, z.Name AS ZoneName,
                   a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblZone z
            JOIN tblArea a  ON z.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            WHERE z.ObjectType = 211
            ORDER BY pa.Name, a.Name, z.Name
        """)

    # Devices — control station devices (for lock state conditions, CT=18, FOT=5)
    devices = q("""
        SELECT csd.ControlStationDeviceID, cs.Name AS DeviceName,
               a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
        FROM tblControlStationDevice csd
        JOIN tblControlStation cs ON csd.ParentControlStationID = cs.ControlStationID
        LEFT JOIN tblArea a  ON cs.ParentID = a.AreaID
        LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        ORDER BY pa.Name, a.Name, cs.Name
    """)

    return jsonify({
        "occupancy": occupancy,
        "cci": cci,
        "cco": cco,
        "areas": areas,
        "zones": zones,
        "shades": shades,
        "sequences": sequences,
        "timeclocks": timeclocks,
        "hvac": hvac,
        "devices": devices,
        "rooms": rooms,
        "integrations": integrations,
        "_errors": errors if errors else None,
    })


@app.route("/api/cond-debug")
def cond_debug():
    """Diagnostic: inspect DB to understand available condition data."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    def q(sql):
        rows = _try_q(sql)
        if not rows or (rows and rows[0].get("__error__")):
            return []
        return [dict(r) for r in rows]

    return jsonify({
        # What ConditionTypes are actually used in this program?
        "eval_types": q("""
            SELECT ConditionType, FirstOperandObjectType, COUNT(*) as cnt
            FROM tblEvaluation WHERE ObjectType IN (237,241)
            GROUP BY ConditionType, FirstOperandObjectType ORDER BY cnt DESC
        """),
        # HVAC zones via dedicated table
        "hvac_zones": q("""
            SELECT z.ZoneID, z.Name AS ZoneName, z.ObjectType,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblZoneHVAC hz
            JOIN tblZone z ON hz.ZoneID = z.ZoneID
            JOIN tblArea a ON z.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            ORDER BY pa.Name, a.Name, z.Name
        """),
        # Zone ObjectTypes distribution — to find CCI/CCO/HVAC patterns
        "zone_object_types": q("""
            SELECT z.ObjectType, COUNT(*) as cnt,
                   MIN(z.Name) as sample_name
            FROM tblZone z GROUP BY z.ObjectType ORDER BY cnt DESC
        """),
        # SwitchLeg LoadTypes — to find CCI/CCO patterns
        "switchleg_loadtypes": q("""
            SELECT sl.LoadType, lt.Description, COUNT(*) as cnt,
                   MIN(z.Name) as sample_zone_name
            FROM tblSwitchLeg sl
            LEFT JOIN lstLoadType lt ON lt.LoadTypeID = sl.LoadType
            LEFT JOIN tblZone z ON z.ZoneID = sl.ParentID
            GROUP BY sl.LoadType, lt.Description ORDER BY sl.LoadType
        """),
        # Enclosures and their sub-tables (for CCI)
        "enclosures": q("""
            SELECT e.EnclosureID, e.Name AS EnclosureName,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblEnclosure e
            LEFT JOIN tblArea a ON e.ParentAreaID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
        """),
        # CCI conditions from existing evaluations
        "cci_evals": q("""
            SELECT e.EvaluationID, e.ConditionType, e.FirstOperandObjectType,
                   e.FirstOperandObjectID, e.FirstOperandRefProperty,
                   e.EvaluationOperator, e.SecondOperand
            FROM tblEvaluation e
            WHERE e.ObjectType IN (237,241)
              AND e.ConditionType NOT IN (0,1,2,3,5,6,7,8,9,10,14,18,23)
        """),
        # CCO switchlegs (LoadType 25=Maintained, 26=Pulsed)
        "cco_switchlegs": q("""
            SELECT sl.SwitchLegID, sl.Name AS SwitchLegName, sl.ParentID,
                   sl.LoadType, lt.Description AS LoadTypeDesc,
                   a.AreaID, a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblSwitchLeg sl
            LEFT JOIN lstLoadType lt ON lt.LoadTypeID = sl.LoadType
            LEFT JOIN tblArea a ON sl.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            WHERE sl.LoadType IN (25, 26)
            ORDER BY pa.Name, a.Name, sl.Name
        """),
        # All ObjectType=15 zones with their SwitchLeg LoadType (to distinguish CCI vs lighting)
        "cci_zone_detail": q("""
            SELECT z.ZoneID, z.Name AS ZoneName, z.ObjectType, z.ParentID AS AreaID,
                   sl.SwitchLegID, sl.LoadType, sl.Name AS SlName,
                   lt.Description AS LoadTypeDesc
            FROM tblZone z
            LEFT JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
            LEFT JOIN lstLoadType lt ON lt.LoadTypeID = sl.LoadType
            WHERE z.ObjectType = 15
            ORDER BY z.ZoneID
        """),
        # Tables related to contact closure / enclosure inputs
        "contact_tables": q("""
            SELECT name FROM sys.tables
            WHERE name LIKE '%Contact%' OR name LIKE '%Enclosure%'
               OR name LIKE '%Input%' OR name LIKE '%CCI%'
            ORDER BY name
        """),
        # All DB tables
        "all_tables": q("SELECT name FROM sys.tables ORDER BY name"),
        # All zone ObjectTypes
        "all_zone_types": q("""
            SELECT z.ObjectType, COUNT(*) as cnt, MIN(z.Name) as sample
            FROM tblZone z GROUP BY z.ObjectType ORDER BY z.ObjectType
        """),
        # Zones in areas that contain QSE-IO enclosures (all ObjectTypes)
        "qseio_area_zones": q("""
            SELECT z.ZoneID, z.Name AS ZoneName, z.ObjectType,
                   sl.LoadType, lt.Description AS LoadTypeDesc,
                   a.Name AS AreaName
            FROM tblZone z
            LEFT JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
            LEFT JOIN lstLoadType lt ON lt.LoadTypeID = sl.LoadType
            JOIN tblArea a ON z.ParentID = a.AreaID
            WHERE EXISTS (
                SELECT 1 FROM tblEnclosure e
                WHERE e.ParentAreaID = a.AreaID
                  AND (e.Name LIKE '%QSE%' OR e.Name LIKE '%IO%')
            )
            ORDER BY a.Name, z.Name
        """),
        # All SensorConnections (full schema + data)
        "cci_sensor_connections": q("SELECT TOP 10 * FROM tblSensorConnection"),
        # SensorConnections linked to doors
        "door_sensor_connections": q("""
            SELECT d.DoorID, d.Name AS DoorName, d.ParentID AS DoorAreaID,
                   d.AssociatedCCISensorConnectionID,
                   sc.Name AS ScName, sc.ObjectType AS ScObjectType
            FROM tblDoor d
            LEFT JOIN tblSensorConnection sc
                   ON sc.SensorConnectionID = d.AssociatedCCISensorConnectionID
        """),
        # Assignment command types by object type — to discover cmd types for timeclock/occupancy/room props
        "assignment_cmd_types": q("""
            SELECT pa.AssignableObjectType, pa.AssignmentCommandType,
                   pa.AssignmentCommandGroup, COUNT(*) as cnt,
                   MIN(pa.AssignableObjectID) as sample_id
            FROM tblPresetAssignment pa
            GROUP BY pa.AssignableObjectType, pa.AssignmentCommandType, pa.AssignmentCommandGroup
            ORDER BY pa.AssignableObjectType, cnt DESC
        """),
        # Assignment params for timeclock (ObjType=19), occupancy (ObjType=38), room props (ObjType=400)
        "target_assignments": q("""
            SELECT TOP 20 pa.PresetAssignmentID, pa.AssignableObjectType,
                   pa.AssignableObjectID, pa.AssignmentCommandType,
                   pa.AssignmentCommandGroup,
                   acp.ParameterType, acp.ParameterValue, acp.SortOrder
            FROM tblPresetAssignment pa
            LEFT JOIN tblAssignmentCommandParameter acp ON acp.ParentId = pa.PresetAssignmentID
            WHERE pa.AssignableObjectType IN (19, 38, 400)
            ORDER BY pa.AssignableObjectType, pa.PresetAssignmentID, acp.SortOrder
        """),
        "shared_scene_area_assignments": q("""
            SELECT TOP 30
                   p.PresetID, p.Name AS PresetName, p.SortOrder AS PresetSortOrder,
                   pa.PresetAssignmentID, pa.AssignmentCommandType, pa.AssignmentCommandGroup,
                   a.AreaID, a.Name AS AreaName,
                   acp.ParameterType, acp.ParameterValue
            FROM tblPreset p
            JOIN tblPresetAssignment pa ON pa.ParentID = p.PresetID
            LEFT JOIN tblArea a ON a.AreaID = pa.AssignableObjectID
            LEFT JOIN tblAssignmentCommandParameter acp ON acp.ParentId = pa.PresetAssignmentID
            WHERE p.PresetType = 3 AND p.ParentType = 2
              AND pa.AssignableObjectType = 2
            ORDER BY p.SortOrder, pa.PresetAssignmentID, acp.ParameterType
        """),
        "shared_scene_zone_assignments": q("""
            SELECT TOP 30
                   p.PresetID, p.Name AS PresetName,
                   pa.PresetAssignmentID, pa.AssignableObjectType, pa.AssignableObjectID,
                   pa.AssignmentCommandType, pa.AssignmentCommandGroup,
                   COALESCE(z15.Name, z198.Name) AS ZoneName,
                   acp.ParameterType, acp.ParameterValue
            FROM tblPreset p
            JOIN tblPresetAssignment pa ON pa.ParentID = p.PresetID
            LEFT JOIN tblZone z15  ON z15.ZoneID  = pa.AssignableObjectID AND pa.AssignableObjectType = 15
            LEFT JOIN tblZone z198 ON z198.ZoneID = pa.AssignableObjectID AND pa.AssignableObjectType = 198
            LEFT JOIN tblAssignmentCommandParameter acp ON acp.ParentId = pa.PresetAssignmentID
            WHERE p.PresetType = 3 AND p.ParentType = 2
              AND pa.AssignableObjectType IN (15, 198, 211)
            ORDER BY p.SortOrder, pa.PresetAssignmentID, acp.ParameterType
        """),
        "hvac_all_preset_params": q("""
            SELECT p.PresetID, p.Name AS PresetName, p.SortOrder,
                   pa.PresetAssignmentID, pa.AssignableObjectID,
                   acp.ParameterType, acp.ParameterValue
            FROM tblPreset p
            JOIN tblPresetAssignment pa ON pa.ParentID = p.PresetID
                 AND pa.AssignableObjectType = 211
            LEFT JOIN tblAssignmentCommandParameter acp ON acp.ParentId = pa.PresetAssignmentID
            WHERE p.PresetType = 3 AND p.ParentType = 2
            ORDER BY p.SortOrder, pa.PresetAssignmentID, acp.ParameterType
        """),
        "area_scenes_sample": q("""
            SELECT TOP 20 sc.ParentID AS AreaID, a.Name AS AreaName,
                   s.SceneID, s.Name AS SceneName, s.Number, s.SortOrder
            FROM tblSceneController sc
            JOIN tblScene s ON s.ParentSceneControllerID = sc.SceneControllerID
            JOIN tblArea a ON a.AreaID = sc.ParentID
            WHERE sc.ParentType = 2
            ORDER BY sc.ParentID, s.Number
        """),
        "scene_template_columns": q("""
            SELECT COLUMN_NAME, DATA_TYPE
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_NAME = 'tblSceneTemplate'
            ORDER BY ORDINAL_POSITION
        """),
        "scene_template_sample": q("SELECT TOP 10 * FROM tblSceneTemplate"),
    })


@app.route("/api/evaluation/<int:eval_id>", methods=["PUT"])
def update_evaluation(eval_id):
    """Update a condition (tblEvaluation row)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    allowed = {"FirstOperandObjectID", "SecondOperand", "FirstOperandRefProperty",
               "EvaluationOperator", "ThirdOperand", "ConditionType", "FirstOperandObjectType",
               "ObjectType"}
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "更新フィールドなし"}), 400
    sets = ", ".join(f"{k} = ?" for k in updates)
    vals = list(updates.values()) + [eval_id]
    try:
        execute_sql(f"UPDATE tblEvaluation SET {sets} WHERE EvaluationID = ?", vals)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/scenes")
def all_scenes():
    """Return callable presets for the scene picker.

    Shared programmable scenes (PresetType=3, ParentType=2) are always returned.
    Button-local action presets (PresetType=1, ParentType=231) are returned only
    when ?pm_id=<id> is provided, filtered to that programming model.
    """
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400

    pm_id = request.args.get("pm_id", type=int)

    # Shared programmable scenes — callable from any button
    shared = _try_q("""
        SELECT p.PresetID, p.Name, p.ParentID AS AreaID,
               a.Name AS AreaName
        FROM tblPreset p
        LEFT JOIN tblArea a ON p.ParentID = a.AreaID
        WHERE p.PresetType = 3 AND p.ParentType = 2
        ORDER BY a.Name, p.Name
    """)
    if shared and shared[0].get("__error__"):
        shared = []

    # Button-local action presets — only for the current PM
    local_presets = []
    if pm_id:
        lp = _try_q("""
            SELECT p.PresetID, p.Name
            FROM tblPreset p
            WHERE p.PresetType = 1 AND p.ParentType = 231 AND p.ParentID = ?
            ORDER BY p.SortOrder, p.Name
        """, (pm_id,))
        if lp and not lp[0].get("__error__"):
            local_presets = lp

    return jsonify({"shared_scenes": shared or [], "local_presets": local_presets})


@app.route("/api/action/<int:action_id>/preset", methods=["PUT"])
def update_action_preset(action_id):
    """Update PresetId for a Run action (ObjectType=234)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    preset_id = data.get("preset_id")  # None → clear
    try:
        execute_sql(
            "UPDATE tblAction SET PresetId = ? WHERE ActionID = ? AND ObjectType = 234",
            (preset_id, action_id)
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/pm/<int:pm_id>", methods=["PUT"])
def update_pm(pm_id):
    """Update ProgrammingModel fields (presets, LED logic, etc.)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    allowed = {
        "PressPresetID", "ReleasePresetID", "HoldPresetId", "DoubleTapPresetID",
        "PresetID", "OnPresetID", "OffPresetID",
        "ReferencePresetIDForLed", "LedLogic", "UseReverseLedLogic",
        "AllowDoubleTap", "HoldTime",
    }
    updates = {k: v for k, v in data.items() if k in allowed}
    if not updates:
        return jsonify({"error": "更新するフィールドがありません"}), 400
    sets = ", ".join(f"[{k}] = ?" for k in updates)
    vals = list(updates.values()) + [pm_id]
    try:
        execute_sql(
            f"UPDATE tblProgrammingModel SET {sets}, NeedsTransfer = 1 WHERE ProgrammingModelID = ?",
            vals
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Button-local Actions (tblPreset PresetType=1) ─────────────────────────────

@app.route("/api/pm/<int:pm_id>/actions", methods=["GET"])
def list_pm_actions(pm_id):
    """List button-local action presets for a ProgrammingModel."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    rows = q("""
        SELECT PresetID, Name, SortOrder
        FROM tblPreset
        WHERE PresetType = 1 AND ParentType = 231 AND ParentID = ?
        ORDER BY SortOrder, Name
    """, (pm_id,))
    return jsonify(rows)


@app.route("/api/pm/<int:pm_id>/actions", methods=["POST"])
def create_pm_action(pm_id):
    """Create a new button-local action preset."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    existing = q("""
        SELECT Name FROM tblPreset
        WHERE PresetType = 1 AND ParentType = 231 AND ParentID = ?
    """, (pm_id,))
    sort_row = q("""
        SELECT ISNULL(MAX(SortOrder), 0) AS max_sort FROM tblPreset
        WHERE PresetType = 1 AND ParentType = 231 AND ParentID = ?
    """, (pm_id,))
    # Auto-increment: find next unused "Action NNN" number
    used_nums = set()
    for r in existing:
        m = re.match(r'^Action (\d+)$', r.get("Name", ""))
        if m:
            used_nums.add(int(m.group(1)))
    n = 1
    while n in used_nums:
        n += 1
    name = f"Action {n:03d}"
    next_sort = (sort_row[0]["max_sort"] if sort_row else 0) + 1
    try:
        preset_id = _alloc_and_insert("tblPreset", "PresetID", {
            "Name": name, "DatabaseRevision": 0, "SortOrder": next_sort,
            "ParentID": pm_id, "ParentType": 231,
            "NeedsTransfer": 1, "PresetType": 1,
            "WhereUsedId": 2147483647, "IsGPDPreset": 0,
        })
        return jsonify({"preset_id": preset_id, "name": name, "sort_order": next_sort})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/preset/<int:preset_id>", methods=["PATCH"])
def update_preset(preset_id):
    """Rename a button-local action preset."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    name = (request.json or {}).get("name", "").strip()
    if not name:
        return jsonify({"error": "名前を入力してください"}), 400
    try:
        old = q("SELECT Name FROM tblPreset WHERE PresetID=?", (preset_id,))
        old_name = old[0]["Name"] if old else ""
        execute_sql("UPDATE tblPreset SET Name = ? WHERE PresetID = ? AND PresetType = 1",
                    (name, preset_id))
        push_undo(
            [("UPDATE tblPreset SET Name=? WHERE PresetID=?", (old_name, preset_id))],
            [("UPDATE tblPreset SET Name=? WHERE PresetID=?", (name, preset_id))],
            f"プリセット名変更 → {name}"
        )
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/preset/<int:preset_id>", methods=["DELETE"])
def delete_preset(preset_id):
    """Delete a button-local action preset and all its assignments."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    try:
        assignments = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID = ?",
                        (preset_id,))
        for a in assignments:
            execute_sql("DELETE FROM tblAssignmentCommandParameter WHERE ParentId = ?",
                        (a["PresetAssignmentID"],))
        execute_sql("DELETE FROM tblPresetAssignment WHERE ParentID = ?", (preset_id,))
        for col in ("PressPresetID", "ReleasePresetID", "HoldPresetId",
                    "DoubleTapPresetID", "PresetID", "ReferencePresetIDForLed"):
            execute_sql(f"UPDATE tblProgrammingModel SET [{col}] = NULL WHERE [{col}] = ?",
                        (preset_id,))
        execute_sql("UPDATE tblAction SET PresetId = 0 WHERE PresetId = ?", (preset_id,))
        execute_sql("DELETE FROM tblPreset WHERE PresetID = ? AND PresetType = 1", (preset_id,))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Assignable items list ──────────────────────────────────────────────────────

@app.route("/api/assignable-items", methods=["GET"])
def get_assignable_items():
    """Get list of assignable items by type."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    item_type = request.args.get("type", "zone")

    if item_type == "zone":
        # Lighting zones: ObjectType=15, excluding CCO (LoadType 25/26)
        return jsonify(q("""
            SELECT z.ZoneID, z.Name, z.ControlType,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblZone z
            LEFT JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
            JOIN tblArea a  ON z.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            WHERE z.ObjectType = 15
              AND (sl.LoadType IS NULL OR sl.LoadType NOT IN (25, 26))
            ORDER BY pa.Name, a.Name, z.Name
        """))

    elif item_type == "cco":
        # Contact Closure Output zones (LoadType 25=Maintained, 26=Pulsed)
        return jsonify(q("""
            SELECT z.ZoneID, z.Name, z.ControlType, sl.LoadType,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblZone z
            JOIN tblSwitchLeg sl ON sl.SwitchLegID = z.ZoneID + 1
            JOIN tblArea a  ON z.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            WHERE sl.LoadType IN (25, 26)
            ORDER BY pa.Name, a.Name, z.Name
        """))

    elif item_type == "shade":
        return jsonify(q("""
            SELECT sg.ShadeGroupID, sg.Name,
                   a.Name AS AreaName
            FROM tblShadeGroup sg
            LEFT JOIN tblArea a ON sg.ParentAreaID = a.AreaID
            ORDER BY a.Name, sg.Name
        """))

    elif item_type == "device":
        return jsonify(q("""
            SELECT csd.ControlStationDeviceID, csd.Name AS DeviceName,
                   cs.Name AS StationName,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblControlStationDevice csd
            JOIN tblControlStation cs ON csd.ParentControlStationID = cs.ControlStationID
            LEFT JOIN tblArea a  ON cs.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            ORDER BY pa.Name, a.Name, cs.Name
        """))

    elif item_type == "variable":
        vars_ = q("SELECT VariableID, Name FROM tblVariable ORDER BY Name")
        for v in vars_:
            v["states"] = q(
                "SELECT VariableStateID, Name FROM tblVariableState WHERE ParentID = ? ORDER BY SortOrder",
                (v["VariableID"],))
        return jsonify(vars_)

    elif item_type == "sequence":
        return jsonify(q("""
            SELECT s.SequenceID, s.Name,
                   a.Name AS AreaName
            FROM tblSequence s
            LEFT JOIN tblArea a ON s.ParentID = a.AreaID AND s.ParentType = 2
            ORDER BY s.Name
        """))

    elif item_type == "hvac":
        return jsonify(q("""
            SELECT h.ZoneID, z.Name,
                   a.Name AS AreaName, pa.Name AS ParentAreaName
            FROM tblZoneHVAC h
            JOIN tblZone z ON z.ZoneID = h.ZoneID
            JOIN tblArea a ON z.ParentID = a.AreaID
            LEFT JOIN tblArea pa ON a.ParentID = pa.AreaID
            ORDER BY pa.Name, a.Name, z.Name
        """))

    else:
        return jsonify({"error": "不明なタイプ"}), 400


# Helper: ControlType → (AssignmentCommandType, AssignmentCommandGroup, level_param_type)
# ControlType: 1=Dimmer, 2=Switched, 5=Shade, 6=Fan, 7=CCO Maintained, 8=CCO Pulsed
_ZONE_CMD_MAP = {
    1: (2,  1, 3),   # Dimmer:       CmdType=2, CmdGroup=1, level=ParamType 3 (0–100)
    2: (2,  1, 3),   # Switched:     same, level is 0 or 100
    5: (2,  1, 3),   # Shade zone:   treated like dimmer
    6: (2,  1, 3),   # Fan:          treated like dimmer
    7: (18, 7, 18),  # CCO Maintained: CmdType=18, CmdGroup=7, level=ParamType 18 (0/100)
    8: (12, 8, None),# CCO Pulsed:   CmdType=12, CmdGroup=8, no level param
}


# ── Preset name update ─────────────────────────────────────────────────────────

@app.route("/api/preset/<int:preset_id>/name", methods=["PUT"])
def update_preset_name(preset_id):
    """Rename a preset (shared scene or button-local)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "名前が空です"}), 400
    old = q("SELECT Name FROM tblPreset WHERE PresetID=?", (preset_id,))
    old_name = old[0]["Name"] if old else ""
    execute_sql("UPDATE tblPreset SET Name=?, NeedsTransfer=1 WHERE PresetID=?", (name, preset_id))
    push_undo(
        [("UPDATE tblPreset SET Name=?, NeedsTransfer=1 WHERE PresetID=?", (old_name, preset_id))],
        [("UPDATE tblPreset SET Name=?, NeedsTransfer=1 WHERE PresetID=?", (name, preset_id))],
        f"シーン名変更 → {name}"
    )
    return jsonify({"ok": True})


# ── Assignments for action presets ─────────────────────────────────────────────

@app.route("/api/preset/<int:preset_id>/assignments", methods=["GET"])
def list_preset_assignments(preset_id):
    """List all assignments for an action preset."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    assignments = q("""
        SELECT pa.PresetAssignmentID, pa.AssignableObjectID, pa.AssignableObjectType,
               pa.AssignmentCommandType, pa.SortOrder,
               -- Zone (ObjType=198 legacy or ObjType=15 direct)
               z198.Name  AS ZoneName198,  z198.ControlType AS CtrlType198,
               z15.Name   AS ZoneName15,   z15.ControlType  AS CtrlType15,
               -- Area (ObjType=2)
               ar.Name    AS AreaName,
               -- Timeclock (ObjType=19)
               tc.Name    AS TimeclockName,
               -- OccupancyGroup (ObjType=38)
               og.Name    AS OccupancyName,
               -- RentableSpace (ObjType=400)
               rs.Name    AS RoomPropName,
               -- IntegrationCommandSet (ObjType=202)
               ics.Name   AS IntegrationName,
               -- Sequence
               seq.Name   AS SeqName,
               -- Variable
               var.Name   AS VarName,
               -- HVAC (via ZoneID=211 in tblZone joined through tblZoneHVAC)
               hz.Name    AS HvacName,
               -- Shade Group
               sg.Name    AS ShadeName,
               -- Device (control station device name)
               cs.Name    AS DeviceName
        FROM tblPresetAssignment pa
        LEFT JOIN tblZone z198    ON pa.AssignableObjectID = z198.ZoneID
                                 AND pa.AssignableObjectType = 198
        LEFT JOIN tblZone z15     ON pa.AssignableObjectID = z15.ZoneID
                                 AND pa.AssignableObjectType = 15
        LEFT JOIN tblArea ar         ON pa.AssignableObjectID = ar.AreaID
                                    AND pa.AssignableObjectType = 2
        LEFT JOIN tblTimeClock tc    ON pa.AssignableObjectID = tc.TimeClockID
                                    AND pa.AssignableObjectType = 19
        LEFT JOIN tblOccupancyGroup og ON pa.AssignableObjectID = og.OccupancyGroupID
                                    AND pa.AssignableObjectType = 38
        LEFT JOIN tblRentableSpace rs  ON pa.AssignableObjectID = rs.RentableSpaceID
                                    AND pa.AssignableObjectType = 400
        LEFT JOIN tblIntegrationCommandSet ics ON pa.AssignableObjectID = ics.IntegrationCommandSetID
                                    AND pa.AssignableObjectType = 202
        LEFT JOIN tblSequence seq    ON pa.AssignableObjectID = seq.SequenceID
                                 AND pa.AssignableObjectType = 90
        LEFT JOIN tblVariable var ON pa.AssignableObjectID = var.VariableID
                                 AND pa.AssignableObjectType = 169
        LEFT JOIN tblZone hz      ON pa.AssignableObjectID = hz.ZoneID
                                 AND pa.AssignableObjectType = 211
        LEFT JOIN tblShadeGroup sg ON pa.AssignableObjectID = sg.ShadeGroupID
                                 AND pa.AssignableObjectType = 133
        LEFT JOIN tblControlStationDevice csd ON pa.AssignableObjectID = csd.ControlStationDeviceID
                                 AND pa.AssignableObjectType = 5
        LEFT JOIN tblControlStation cs ON csd.ParentControlStationID = cs.ControlStationID
        WHERE pa.ParentID = ?
        ORDER BY pa.SortOrder
    """, (preset_id,))

    for a in assignments:
        ot = a["AssignableObjectType"]
        ct = a["AssignmentCommandType"]

        # Consolidate zone name and ControlType
        if ot == 198:
            a["ZoneName"] = a["ZoneName198"]
            a["ControlType"] = a["CtrlType198"]
        elif ot == 15:
            a["ZoneName"] = a["ZoneName15"]
            a["ControlType"] = a["CtrlType15"]
        else:
            a["ZoneName"] = None
            a["ControlType"] = None
        # Remove redundant fields
        for k in ("ZoneName198","CtrlType198","ZoneName15","CtrlType15"):
            a.pop(k, None)

        # Consolidate display names for new types
        if ot != 2:
            a.pop("AreaName", None)
        if ot != 19:
            a.pop("TimeclockName", None)
        if ot != 38:
            a.pop("OccupancyName", None)
        if ot != 400:
            a.pop("RoomPropName", None)
        if ot != 202:
            a.pop("IntegrationName", None)

        # Load raw parameters
        params = q("""
            SELECT ParameterType, ParameterValue
            FROM tblAssignmentCommandParameter WHERE ParentId = ? ORDER BY SortOrder
        """, (a["PresetAssignmentID"],))
        pm = {p["ParameterType"]: p["ParameterValue"] for p in params}
        a["_params"] = pm   # keep raw for debugging

        a["fade"]  = round(pm.get(1, 0) / 4, 2)  # Lutron stores in 250ms units; convert to seconds
        a["delay"] = round(pm.get(2, 0) / 4, 2)

        if ot == 198:
            # Legacy: level stored 0–10000, normalise to 0–100 for UI
            raw = pm.get(17, 10000)
            a["level"] = raw // 100
        elif ot == 15:
            if ct == 2:
                a["level"] = pm.get(3, 100)     # Dimmer/Switched 0–100
            elif ct == 18:
                a["level"] = pm.get(18, 100)    # CCO Maintained 0/100
            else:
                a["level"] = None               # CCO Pulsed – no level
        elif ot == 2:
            if ct == 5:   # Shared Scene recall: scene Number stored in PT=7
                a["level"] = pm.get(7)
            else:         # Regular area level: PT=3, 0–100
                a["level"] = pm.get(3, 100)
        elif ot == 38:
            a["level"] = pm.get(35, 1)          # Occupancy state 1/2/4
        elif ot == 400:
            # Confirmed from scene names in LD DB:
            # "Check in room"=PT76:0, "Check Out Room"=PT76:1 → PT76=RoomState(152) 0=CheckIn,1=CheckOut
            # "On/Off Level" PT77=1/0 → PT77=Presence(149) 1=On,0=Off
            # "First key->True" PT78=1 → PT78=FGE(154) 1=True,0=False
            # "TURNDOWN ON" PT79=1 → PT79=Turndown(155) 1=Enabled,0=Disabled
            # "DND ON/OFF" PT80=1/0 → PT80=DND(151) 1=Enabled,0=Disabled
            # "MUR ON/OFF" PT81=1/0 → PT81=MUR(153) 1=Enabled,0=Disabled
            # PT84=Automation(157) 1=Enabled,0=Disabled
            # Absent row OR value=65280 → Unaffected (null)
            def _rp(v): return None if (v is None or v == 65280) else v
            a["props"] = {
                152: _rp(pm.get(76)),  # RoomState
                149: _rp(pm.get(77)),  # Presence
                154: _rp(pm.get(78)),  # FGE
                155: _rp(pm.get(79)),  # Turndown
                151: _rp(pm.get(80)),  # DND
                153: _rp(pm.get(81)),  # MUR
                157: _rp(pm.get(84)),  # Automation
            }
            a["level"] = None
        elif ot == 211:
            # HVAC programming params (confirmed from DB analysis):
            # PT=44: schedule override (255=Unaffected→0, 1=Run, 2=Hold confirmed from DB)
            # PT=47: op mode (255=Unaffected, 1=Heat, 2=Cool, 3=Auto, 4=Off)
            # PT=48: fan speed (255=Unaffected, 0=Unaffected(alt), 1=Auto, 2=High, 3=Med, 4=Low)
            # PT=53: "temp setpoint active" flag (NOT Schedule) — 1 when PT=54 is set, auto-managed
            # PT=54: setpoint in tenths of °F (720=Unaffected; 662=19°C, 716=22°C)
            # PT=55: below-setpoint (heat) drift in tenths of °F (0=none, -18=-1°C, -36=-2°C)
            # PT=56: above-setpoint (cool) drift in tenths of °F (0=none, 18=+1°C, 36=+2°C)
            # PT=57: drift enable flag (0=drift not applied, 1=drift applied)
            mode_raw  = pm.get(47, 255)
            fan_raw   = pm.get(48, 255)
            sched_raw = pm.get(44, 255)
            a["hvac"] = {
                "setpoint":   pm.get(54, 720),
                "schedule":   0 if sched_raw == 255 else sched_raw,  # normalize 255→0
                "mode":       0 if mode_raw  == 255 else mode_raw,   # normalize 255→0
                "fan":        0 if fan_raw   == 255 else fan_raw,    # normalize 255→0
                "heat_drift": pm.get(55, 0),   # tenths of °F, negative (÷18 = °C)
                "cool_drift": pm.get(56, 0),   # tenths of °F, positive (÷18 = °C)
                "drift_on":   pm.get(57, 0),   # 1 when drift settings are active
            }
            a["level"] = None
        elif ot == 133:
            a["level"] = pm.get(7, 0)           # Shade position 0–100
        elif ot == 5:
            a["level"] = pm.get(22, 0)          # Device: lock value
        else:
            a["level"] = None

        # Variable: target state
        a["state_id"] = pm.get(32)
        if ot == 169 and a["state_id"]:
            rows = q("SELECT Name FROM tblVariableState WHERE VariableStateID = ?",
                     (a["state_id"],))
            a["StateName"] = rows[0]["Name"] if rows else None
        else:
            a["StateName"] = None

    # If ?shared=1, enrich area assignments with scene info from tblScene
    if request.args.get("shared") == "1":
        prows = _try_q("SELECT SortOrder FROM tblPreset WHERE PresetID = ?", (preset_id,))
        scene_num = prows[0]["SortOrder"] if prows else None
        for a in assignments:
            if a.get("AssignableObjectType") == 2:
                area_id = a["AssignableObjectID"]
                area_scenes = _try_q("""
                    SELECT s.SceneID, s.Name, s.Number
                    FROM tblSceneController sc
                    JOIN tblScene s ON s.ParentSceneControllerID = sc.SceneControllerID
                    WHERE sc.ParentID = ? AND sc.ParentType = 2
                    ORDER BY s.Number
                """, (area_id,))
                a["area_scenes"] = area_scenes
                # PT=3 stores the scene number to recall; fall back to preset SortOrder
                area_scene_num = a.get("level") if a.get("level") is not None else scene_num
                matched = [s for s in area_scenes if s["Number"] == area_scene_num]
                a["area_scene_name"] = matched[0]["Name"] if matched else None
                # Ensure level reflects the resolved scene number so the dropdown selects correctly
                a["level"] = area_scene_num

    return jsonify(assignments)


@app.route("/api/preset/<int:preset_id>/assignments", methods=["POST"])
def add_preset_assignment(preset_id):
    """Add an item assignment to an action preset."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}
    item_type = data.get("item_type", "zone")

    if item_type in ("zone", "cco"):
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        # Fetch ControlType to choose the right CmdType/CmdGroup/param
        zrow = q("SELECT ControlType FROM tblZone WHERE ZoneID = ?", (item_id,))
        ctrl = zrow[0]["ControlType"] if zrow else 1
        cmd_type, cmd_group, lv_param = _ZONE_CMD_MAP.get(ctrl, (2, 1, 3))
        obj_type = 15
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=15 AND AssignmentCommandType=?",
                     (preset_id, item_id, cmd_type))
        if existing:
            return jsonify({"error": "このゾーンはすでに割り当て済みです"}), 400
        level = int(data.get("level", 100))
        fade  = round(float(data.get("fade",  0)) * 4)  # seconds → 250ms units
        delay = round(float(data.get("delay", 0)) * 4)
        if lv_param is not None:
            params = [(0, 1, fade), (1, 2, delay), (2, lv_param, level)]
        else:
            params = [(0, 1, fade), (1, 2, delay)]  # CCO Pulsed: no level, but has pulse duration

    elif item_type == "shade":
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 133, 5, 11
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=133",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このシェードグループはすでに割り当て済みです"}), 400
        pos = int(data.get("level", 0))
        params = [(0, 2, round(float(data.get("delay", 0)) * 4)), (1, 7, pos)]

    elif item_type == "device":
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 5, 19, 20
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=5",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このデバイスはすでに割り当て済みです"}), 400
        lock_val = int(data.get("level", 1))  # 1=Lock, 2=Unlock
        params = [(0, 22, lock_val), (1, 29, 0), (2, 2, round(float(data.get("delay", 0)) * 4))]

    elif item_type == "sequence":
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_group = 90, 16
        cmd_type = int(data.get("cmd_type", 21))  # 21=Start, 23=Pause
        if cmd_type not in (21, 23):
            return jsonify({"error": "cmd_type は 21(開始) または 23(一時停止) のみ有効"}), 400
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=90 AND AssignmentCommandType=?",
                     (preset_id, item_id, cmd_type))
        if existing:
            return jsonify({"error": "このシーケンスはすでに割り当て済みです"}), 400
        params = [(0, 2, round(float(data.get("delay", 0)) * 4))]

    elif item_type == "variable":
        item_id  = data.get("item_id")
        state_id = data.get("state_id")
        if item_id is None or state_id is None:
            return jsonify({"error": "item_id と state_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 169, 39, 27
        existing = q("""SELECT pa.PresetAssignmentID FROM tblPresetAssignment pa
                        JOIN tblAssignmentCommandParameter acp ON acp.ParentId=pa.PresetAssignmentID
                        WHERE pa.ParentID=? AND pa.AssignableObjectID=? AND pa.AssignableObjectType=169
                          AND acp.ParameterType=32 AND acp.ParameterValue=?""",
                     (preset_id, item_id, state_id))
        if existing:
            return jsonify({"error": "この変数とステートの組み合わせはすでに割り当て済みです"}), 400
        params = [(0, 32, int(state_id)), (1, 2, round(float(data.get("delay", 0)) * 4))]

    elif item_type == "area":
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=2",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このエリアはすでに割り当て済みです"}), 400
        # Shared Scenes (PresetType=3) use cmd_type=5/group=3 and store scene# in PT=7
        prow = q("SELECT PresetType FROM tblPreset WHERE PresetID=?", (preset_id,))
        is_shared = prow and prow[0]["PresetType"] == 3
        if is_shared:
            obj_type, cmd_type, cmd_group = 2, 5, 3
            scene_num = int(data.get("level", 1))  # default to scene 1
            delay = round(float(data.get("delay", 0)) * 4)
            params = [(0, 2, delay), (1, 7, scene_num)]
        else:
            obj_type, cmd_type, cmd_group = 2, 2, 1
            level = int(data.get("level", 100))
            fade  = round(float(data.get("fade",  0)) * 4)
            delay = round(float(data.get("delay", 0)) * 4)
            params = [(0, 1, fade), (1, 2, delay), (2, 3, level)]

    elif item_type == "hvac":
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 211, 59, 31
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=211",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このHVACゾーンはすでに割り当て済みです"}), 400
        params = [(0, 44, 255), (1, 47, 255), (2, 48, 255),
                  (3, 53, 0), (4, 54, 720), (5, 55, 0), (6, 56, 0), (7, 57, 0)]

    elif item_type == "timeclock":
        # NOTE: cmd_type/group/params need verification via /api/cond-debug assignment_cmd_types
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        enable = int(data.get("cmd_type", 1))  # 1=Enable, 0=Disable
        obj_type, cmd_type, cmd_group = 19, (36 if enable else 37), 17
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=19 AND AssignmentCommandType=?",
                     (preset_id, item_id, cmd_type))
        if existing:
            return jsonify({"error": "このタイムクロックはすでに割り当て済みです"}), 400
        delay = round(float(data.get("delay", 0)) * 4)
        params = [(0, 2, delay)]

    elif item_type == "occupancy":
        # NOTE: cmd_type/group/params need verification via /api/cond-debug assignment_cmd_types
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 38, 46, 22
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=38",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このOccupancy Groupはすでに割り当て済みです"}), 400
        state_val = int(data.get("level", 1))  # 1=Occupied, 2=Unoccupied, 4=Bypass
        delay = round(float(data.get("delay", 0)) * 4)
        params = [(0, 35, state_val), (1, 2, delay)]

    elif item_type == "roomprop":
        # Confirmed from LD scene names: ObjType=400, CmdType=80, CmdGroup=47
        # PT76=RoomState(152) 0=CheckIn,1=CheckOut
        # PT77=Presence(149) 1=On,0=Off
        # PT78=FGE(154) 1=True,0=False
        # PT79=Turndown(155) 1=Enabled,0=Disabled
        # PT80=DND(151) 1=Enabled,0=Disabled
        # PT81=MUR(153) 1=Enabled,0=Disabled
        # PT84=Automation(157) 1=Enabled,0=Disabled
        item_id = int(data.get("item_id", 34))
        obj_type, cmd_type, cmd_group = 400, 80, 47
        PROP_TO_PT = {152: 76, 149: 77, 154: 78, 155: 79, 151: 80, 153: 81, 157: 84}
        props = data.get("props", {})  # only explicitly set values; absent key = Unaffected
        params = []
        for sort_i, (prop_id_int, pt) in enumerate(sorted(PROP_TO_PT.items(), key=lambda x: x[1])):
            val = props.get(str(prop_id_int))
            if val is not None:  # None = Unaffected, skip (no row)
                params.append((sort_i, pt, int(val)))

    elif item_type == "integration":
        # ObjType=202 (IntegrationCommandSet), confirmed CmdType=52, CmdGroup=36
        item_id = data.get("item_id")
        if item_id is None:
            return jsonify({"error": "item_id が必要です"}), 400
        obj_type, cmd_type, cmd_group = 202, 52, 36
        existing = q("SELECT PresetAssignmentID FROM tblPresetAssignment WHERE ParentID=? AND AssignableObjectID=? AND AssignableObjectType=202",
                     (preset_id, item_id))
        if existing:
            return jsonify({"error": "このIntegrationコマンドはすでに割り当て済みです"}), 400
        delay = int(data.get("delay", 0))
        params = [(0, 2, delay)]

    else:
        return jsonify({"error": f"不明なタイプ: {item_type}"}), 400

    try:
        sort = q("SELECT ISNULL(MAX(SortOrder), -1) + 1 AS s FROM tblPresetAssignment WHERE ParentID = ?",
                 (preset_id,))[0]["s"]
        aid = _alloc_and_insert("tblPresetAssignment", "PresetAssignmentID", {
            "Name": "", "DatabaseRevision": 0, "SortOrder": sort,
            "ParentID": preset_id, "ParentType": 43,
            "AssignableObjectID": item_id, "AssignableObjectType": obj_type,
            "AssignmentCommandType": cmd_type, "NeedsTransfer": 1,
            "AssignmentCommandGroup": cmd_group, "WhereUsedId": 2147483647,
        })
        for sort_p, ptype, pval in params:
            execute_sql(
                "INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (?,?,?,?)",
                (sort_p, aid, ptype, pval))
        # Undo: delete the created assignment + params
        undo_sqls = [
            ("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=?", (aid,)),
            ("DELETE FROM tblPresetAssignment WHERE PresetAssignmentID=?", (aid,)),
        ]
        redo_sqls = [
            ("INSERT INTO tblPresetAssignment (PresetAssignmentID,Name,DatabaseRevision,SortOrder,ParentID,ParentType,AssignableObjectID,AssignableObjectType,AssignmentCommandType,NeedsTransfer,AssignmentCommandGroup,WhereUsedId) VALUES (?,?,0,?,?,43,?,?,?,1,?,2147483647)",
             (aid, "", sort, preset_id, item_id, obj_type, cmd_type, cmd_group)),
        ] + [("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (?,?,?,?)",
              (sort_p, aid, ptype, pval)) for sort_p, ptype, pval in params]
        push_undo(undo_sqls, redo_sqls, "アサイン追加")
        return jsonify({"assignment_id": aid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/assignment/<int:assignment_id>", methods=["PATCH"])
def update_assignment(assignment_id):
    """Update parameters for an assignment."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    data = request.json or {}

    # Unaffected: set cmd_type=1, remove all parameters
    if data.get("unaffected"):
        try:
            old_a = q("SELECT AssignmentCommandType FROM tblPresetAssignment WHERE PresetAssignmentID=?",
                      (assignment_id,))
            old_params = q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=?",
                           (assignment_id,))
            old_ct = old_a[0]["AssignmentCommandType"] if old_a else 1
            undo_sqls = [
                ("UPDATE tblPresetAssignment SET AssignmentCommandType=?, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                 (old_ct, assignment_id)),
                ("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=?", (assignment_id,)),
            ] + [("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (0,?,?,?)",
                  (assignment_id, p["ParameterType"], p["ParameterValue"])) for p in old_params]
            redo_sqls = [
                ("UPDATE tblPresetAssignment SET AssignmentCommandType=1, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                 (assignment_id,)),
                ("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=?", (assignment_id,)),
            ]
            execute_sql("UPDATE tblPresetAssignment SET AssignmentCommandType=1, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                        (assignment_id,))
            execute_sql("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=?",
                        (assignment_id,))
            push_undo(undo_sqls, redo_sqls, "Unaffected設定")
            return jsonify({"ok": True})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # Update AssignmentCommandType if requested (e.g. sequence Start→Pause)
    if "cmd_type" in data:
        try:
            execute_sql("UPDATE tblPresetAssignment SET AssignmentCommandType=?, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                        (int(data["cmd_type"]), assignment_id))
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    # Build param-type → value mapping
    mapping = {}
    if "fade"  in data: mapping[1]  = round(float(data["fade"]) * 4)   # seconds → 250ms units
    if "delay" in data: mapping[2]  = round(float(data["delay"]) * 4)
    if "state_id" in data: mapping[32] = int(data["state_id"])

    if "level" in data:
        # Look up assignment to determine correct ParameterType for level
        assn = q("SELECT AssignableObjectType, AssignmentCommandType FROM tblPresetAssignment WHERE PresetAssignmentID=?",
                 (assignment_id,))
        if assn:
            ot = assn[0]["AssignableObjectType"]
            ct = assn[0]["AssignmentCommandType"]
            if ot == 198:
                mapping[17] = int(data["level"]) * 100  # legacy: 0–10000 scale
            elif ot == 15:
                if ct == 2:
                    mapping[3]  = int(data["level"])   # Dimmer/Switched: ParamType 3, 0–100
                elif ct == 18:
                    mapping[18] = int(data["level"])   # CCO Maintained: ParamType 18, 0/100
            elif ot == 2:
                if ct == 5:  # Shared Scene recall: scene Number in PT=7
                    mapping[7] = int(data["level"])
                else:        # Regular area level: PT=3, 0–100
                    mapping[3] = int(data["level"])
            elif ot == 38:
                mapping[35] = int(data["level"])       # Occupancy state: ParamType 35
            # Room property: handled separately via "props" field, not "level"
            elif ot == 133:
                mapping[7]  = int(data["level"])       # Shade position: ParamType 7
            elif ot == 5:
                mapping[22] = int(data["level"])       # Device lock: ParamType 22

    # HVAC params dict (ObjType=211)
    if "hvac" in data and isinstance(data["hvac"], dict):
        HVAC_PT = {"heat_drift": 55, "cool_drift": 56, "drift_on": 57}
        for key, pt in HVAC_PT.items():
            if key in data["hvac"]:
                mapping[pt] = int(data["hvac"][key])
        # setpoint: also auto-set PT=53 ("temp active" flag)
        if "setpoint" in data["hvac"]:
            sp = int(data["hvac"]["setpoint"])
            mapping[54] = sp
            mapping[53] = 0 if sp == 720 else 1
        # schedule/mode/fan: UI sends 0 for Unaffected; LD stores 255 for Unaffected
        if "schedule" in data["hvac"]:
            mapping[44] = 255 if int(data["hvac"]["schedule"]) == 0 else int(data["hvac"]["schedule"])
        if "mode" in data["hvac"]:
            mapping[47] = 255 if int(data["hvac"]["mode"]) == 0 else int(data["hvac"]["mode"])
        if "fan" in data["hvac"]:
            mapping[48] = 255 if int(data["hvac"]["fan"]) == 0 else int(data["hvac"]["fan"])

    # Room property: None/null → DELETE row (Unaffected = no row)
    # PT76=RoomState(152), PT77=Presence(149), PT78=FGE(154), PT79=Turndown(155)
    # PT80=DND(151), PT81=MUR(153), PT84=Automation(157)
    roomprop_deletes = []
    if "props" in data and isinstance(data["props"], dict):
        PROP_TO_PT = {152: 76, 149: 77, 154: 78, 155: 79, 151: 80, 153: 81, 157: 84}
        for prop_id_str, val in data["props"].items():
            try:
                prop_id = int(prop_id_str)
            except (ValueError, TypeError):
                continue
            if prop_id not in PROP_TO_PT:
                continue
            pt = PROP_TO_PT[prop_id]
            if val is None:  # Unaffected: delete PT row
                roomprop_deletes.append(pt)
            else:
                mapping[pt] = int(val)

    try:
        # Capture before-state for undo (only level/fade/delay for simplicity)
        _undo_ptypes = set(mapping.keys()) | {p for p in roomprop_deletes}
        _old_params = {}
        if _undo_ptypes:
            for row in q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=?",
                         (assignment_id,)):
                _old_params[row["ParameterType"]] = row["ParameterValue"]
        _old_ct_row = q("SELECT AssignmentCommandType FROM tblPresetAssignment WHERE PresetAssignmentID=?",
                        (assignment_id,)) if "cmd_type" in data else []
        _old_ct = _old_ct_row[0]["AssignmentCommandType"] if _old_ct_row else None

        for ptype in roomprop_deletes:
            execute_sql("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=? AND ParameterType=?",
                        (assignment_id, ptype))
        for ptype, pval in mapping.items():
            rows = q("SELECT 1 FROM tblAssignmentCommandParameter WHERE ParentId=? AND ParameterType=?",
                     (assignment_id, ptype))
            if rows:
                execute_sql("UPDATE tblAssignmentCommandParameter SET ParameterValue=? WHERE ParentId=? AND ParameterType=?",
                            (pval, assignment_id, ptype))
            else:
                execute_sql("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (?,?,?,?)",
                            (0, assignment_id, ptype, pval))

        # Build undo/redo for level/fade/delay/cmd_type changes
        if _undo_ptypes or _old_ct is not None:
            undo_sqls, redo_sqls = [], []
            if _old_ct is not None:
                undo_sqls.append(("UPDATE tblPresetAssignment SET AssignmentCommandType=?, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                                  (_old_ct, assignment_id)))
                redo_sqls.append(("UPDATE tblPresetAssignment SET AssignmentCommandType=?, NeedsTransfer=1 WHERE PresetAssignmentID=?",
                                  (int(data["cmd_type"]), assignment_id)))
            for pt in _undo_ptypes:
                if pt in _old_params:
                    undo_sqls.append(("UPDATE tblAssignmentCommandParameter SET ParameterValue=? WHERE ParentId=? AND ParameterType=?",
                                      (_old_params[pt], assignment_id, pt)))
                else:
                    undo_sqls.append(("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=? AND ParameterType=?",
                                      (assignment_id, pt)))
                if pt in mapping:
                    redo_sqls.append(("UPDATE tblAssignmentCommandParameter SET ParameterValue=? WHERE ParentId=? AND ParameterType=?",
                                      (mapping[pt], assignment_id, pt)))
                else:
                    redo_sqls.append(("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=? AND ParameterType=?",
                                      (assignment_id, pt)))
            label = f"レベル変更 (ID:{assignment_id})"
            if "level" in data:
                label = f"レベル {data['level']}%"
            push_undo(undo_sqls, redo_sqls, label)

        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/assignment/<int:assignment_id>", methods=["DELETE"])
def delete_assignment(assignment_id):
    """Delete a zone assignment from an action preset."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    try:
        old_a = q("SELECT * FROM tblPresetAssignment WHERE PresetAssignmentID=?", (assignment_id,))
        old_params = q("SELECT ParameterType, ParameterValue FROM tblAssignmentCommandParameter WHERE ParentId=?",
                       (assignment_id,))
        execute_sql("DELETE FROM tblAssignmentCommandParameter WHERE ParentId = ?", (assignment_id,))
        execute_sql("DELETE FROM tblPresetAssignment WHERE PresetAssignmentID = ?", (assignment_id,))
        if old_a:
            a = old_a[0]
            undo_sqls = [
                ("INSERT INTO tblPresetAssignment (PresetAssignmentID,Name,DatabaseRevision,SortOrder,ParentID,ParentType,AssignableObjectID,AssignableObjectType,AssignmentCommandType,NeedsTransfer,AssignmentCommandGroup,WhereUsedId) VALUES (?,?,?,?,?,?,?,?,?,1,?,2147483647)",
                 (assignment_id, a.get("Name",""), a.get("DatabaseRevision",0), a.get("SortOrder",0),
                  a["ParentID"], a["ParentType"], a["AssignableObjectID"], a["AssignableObjectType"],
                  a["AssignmentCommandType"], a.get("AssignmentCommandGroup",1))),
            ] + [("INSERT INTO tblAssignmentCommandParameter (SortOrder,ParentId,ParameterType,ParameterValue) VALUES (0,?,?,?)",
                  (assignment_id, p["ParameterType"], p["ParameterValue"])) for p in old_params]
            redo_sqls = [
                ("DELETE FROM tblAssignmentCommandParameter WHERE ParentId=?", (assignment_id,)),
                ("DELETE FROM tblPresetAssignment WHERE PresetAssignmentID=?", (assignment_id,)),
            ]
            push_undo(undo_sqls, redo_sqls, "アサイン削除")
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/action/<int:action_id>/execution-type", methods=["PATCH"])
def update_action_execution_type(action_id):
    """Update ExecutionType for a Run action (1=Activate,2=Raise,3=Lower,4=Stop)."""
    if not state["db_name"]:
        return jsonify({"error": "DB未接続"}), 400
    exec_type = (request.json or {}).get("execution_type")
    if exec_type not in (1, 2, 3, 4):
        return jsonify({"error": "無効なexecution_type"}), 400
    try:
        execute_sql(
            "UPDATE tblAction SET ExecutionType = ? WHERE ActionID = ? AND ObjectType = 234",
            (exec_type, action_id))
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Save back to .pl ──────────────────────────

@app.route("/api/save", methods=["POST"])
def save():
    if not state["pl_path"]:
        return jsonify({"error": "フォルダモードでは保存できません（.plファイルを開いてください）"}), 400
    try:
        if state["dirty"]:
            save_back_to_pl()
        filename = state.get("original_filename", "project.pl")
        return send_file(state["pl_path"], as_attachment=True, download_name=filename)
    except Exception as e:
        app.logger.error(f"[save] Error: {e}", exc_info=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/shutdown", methods=["POST"])
def shutdown():
    if state["db_name"]:
        drop_db(state["db_name"])
    os.kill(os.getpid(), 9)
    return jsonify({"ok": True})


# ── Undo / Redo ───────────────────────────────

@app.route("/api/undo-status")
def undo_status():
    us = state["undo_stack"]
    rs = state["redo_stack"]
    return jsonify({
        "can_undo": len(us) > 0,
        "can_redo": len(rs) > 0,
        "undo_label": us[-1]["label"] if us else "",
        "redo_label": rs[-1]["label"] if rs else "",
    })


@app.route("/api/undo", methods=["POST"])
def do_undo():
    if not state["undo_stack"]:
        return jsonify({"error": "元に戻す操作がありません"}), 400
    op = state["undo_stack"].pop()
    try:
        execute_sqls(op["sqls"])
        state["redo_stack"].append(op)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "label": op["label"]})


@app.route("/api/redo", methods=["POST"])
def do_redo():
    if not state["redo_stack"]:
        return jsonify({"error": "やり直す操作がありません"}), 400
    op = state["redo_stack"].pop()
    try:
        execute_sqls(op["redo_sqls"])
        state["undo_stack"].append(op)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"ok": True, "label": op["label"]})


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def open_browser():
    import time
    time.sleep(1)
    webbrowser.open("http://127.0.0.1:5000")


if __name__ == "__main__":
    t = threading.Thread(target=open_browser, daemon=True)
    t.start()
    app.run(debug=False, port=5000)
